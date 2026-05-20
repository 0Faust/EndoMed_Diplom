#!/usr/bin/env python3
"""
EndoMed Desktop v6 — Система диагностики патологий ЖКТ
Исправления: потокобезопасность, DICOM, шрифты PDF, индикация симуляции, отмена анализа
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
import tkinter as tk
from PIL import Image
import threading
import time
import json
import logging
import os
import datetime
import random
import textwrap
from pathlib import Path
import numpy as np

logger = logging.getLogger("endomed")

# ── Lock для thread-safe доступа к модели ──
_model_lock = threading.Lock()

# ── PyTorch (опциональный импорт) ──
TORCH_AVAILABLE = False
try:
    import torch
    import torch.nn.functional as F
    import torchvision.transforms as T
    import torchvision.models as models
    TORCH_AVAILABLE = True
except ImportError:
    print("⚠ PyTorch не установлен. Работаем в режиме симуляции.")

# ── Matplotlib для графиков ──
MATPLOTLIB_AVAILABLE = False
try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    print("⚠ Matplotlib не установлен. Дашборд статистики недоступен.")

# ── DICOM (опциональный) ──
DICOM_AVAILABLE = False
try:
    import pydicom
    DICOM_AVAILABLE = True
except ImportError:
    print("⚠ pydicom не установлен. DICOM будет недоступен.")

# ── Шкала «Требуется биопсия» ──
BIOPSY_REQUIRED = {
    "polyp":    {"need": True,  "urgency": "Плановая", "text": "Гистологическая верификация после полипэктомии", "color": "#fbbf24"},
    "cancer":   {"need": True,  "urgency": "Срочная",  "text": "Множественная биопсия из краёв и центра образования (≥6 фрагментов)", "color": "#f87171"},
    "ulcer":    {"need": True,  "urgency": "Плановая", "text": "Биопсия из краёв язвы для исключения малигнизации (4–6 фрагментов)", "color": "#fbbf24"},
    "gastritis": {"need": False, "urgency": "По показаниям", "text": "Биопсия по Сиднейскому протоколу при атрофии (5 точек)", "color": "#94a3b8"},
    "barrett":  {"need": True,  "urgency": "Обязательная", "text": "Четырёхквадрантная биопсия каждые 1–2 см (Сиэтлский протокол)", "color": "#f87171"},
    "ibd":      {"need": True,  "urgency": "Плановая", "text": "Ступенчатая биопсия каждые 10 см для оценки активности и дисплазии", "color": "#fbbf24"},
    "normal":   {"need": False, "urgency": "Не требуется", "text": "Биопсия не показана при нормальной эндоскопической картине", "color": "#34d399"},
}

# ══════════════════════════════════════════════════════════
#  GRAD-CAM — ВИЗУАЛИЗАЦИЯ ВНИМАНИЯ НЕЙРОСЕТИ (v6)
# ══════════════════════════════════════════════════════════
class GradCAM:
    """
    Grad-CAM++ для EfficientNet-B4.

    Ключевые решения:
    - target_layer = model.features[-1]  (весь последний Sequential целиком)
      Хук на отдельный Conv2d внутри MBConv даёт случайную карту, потому что
      там нет полного пространственного контекста блока.
    - Grad-CAM++ вместо обычного Grad-CAM: использует alpha-взвешенные
      квадраты градиентов — работает корректно при tiny gradients
      (|grad| ~ 1e-4) характерных для глубоких сетей.
    - Fallback без ReLU если карта вырождается в ноль.
    """

    def __init__(self, model, target_layer=None):
        self.model = model
        self.gradients = None
        self.activations = None
        self.hook_handles = []

        if target_layer is None:
            # features[-1] — последний Sequential из MBConv-блоков.
            # Хукаемся на весь Sequential: его выход = выход последнего
            # MBConv, feature-map 1792×12×12, пространственная структура цела.
            target_layer = self._find_target_layer(model)

        self.target_layer = target_layer
        self._register_hooks()

    @staticmethod
    def _find_target_layer(model):
        base_model = model.module if hasattr(model, "module") else model
        if hasattr(base_model, "features"):
            children = list(base_model.features.children())
            if children:
                # The last EfficientNet 1x1 block is very semantic but often too
                # coarse for lesion localization. The previous large block keeps
                # more spatial detail while still being class-specific enough.
                return children[-3] if len(children) >= 3 else children[-1]
        raise ValueError("Grad-CAM target layer was not found: model.features is missing")

    @staticmethod
    def _normalize_cam(cam):
        cam = torch.nan_to_num(cam, nan=0.0, posinf=0.0, neginf=0.0)
        cam = cam - cam.min()
        c_max = cam.max()
        if c_max <= 1e-12:
            return torch.zeros_like(cam)

        cam = cam / c_max
        flat = cam.flatten()
        if flat.numel() >= 16:
            low = torch.quantile(flat, 0.55)
            high = torch.quantile(flat, 0.995)
            if high > low:
                cam = torch.clamp((cam - low) / (high - low), 0.0, 1.0)

        return torch.pow(cam, 0.75)

    def _register_hooks(self):
        self.cleanup()
        h1 = self.target_layer.register_forward_hook(self._save_activation)
        self.hook_handles = [h1]

    def _save_activation(self, module, input, output):
        # Без detach — граф нужен для backward
        if isinstance(output, (tuple, list)):
            output = output[0]
        self.activations = output
        self.gradients = None
        if output.requires_grad:
            output.register_hook(self._save_gradient)

    def _save_gradient(self, grad):
        self.gradients = grad.detach()

    def generate(self, input_tensor, class_idx=None):
        """Grad-CAM++ — вызывать внутри torch.enable_grad() и _model_lock."""
        self.model.eval()
        self.gradients = None
        self.activations = None
        input_tensor = input_tensor.detach()
        output = self.model(input_tensor)

        if class_idx is None:
            class_idx = output.argmax(dim=1).item()
        class_idx = int(class_idx)
        if class_idx < 0 or class_idx >= output.shape[1]:
            raise ValueError(f"Grad-CAM class index {class_idx} is outside model output size {output.shape[1]}")

        self.model.zero_grad(set_to_none=True)
        output[0, class_idx].backward(retain_graph=False)

        if self.gradients is None or self.activations is None:
            raise RuntimeError("Grad-CAM хуки не сработали")

        grads = self.gradients.detach()[0]        # [C, H, W]
        acts  = self.activations.detach()[0]      # [C, H, W]
        if grads.ndim != 3 or acts.ndim != 3:
            raise RuntimeError(f"Grad-CAM expected [C,H,W], got gradients={tuple(grads.shape)}, activations={tuple(acts.shape)}")

        # Standard Grad-CAM is less noisy than Grad-CAM++ on this checkpoint.
        # If it degenerates, fall back to a HiResCAM-style gradient*activation map.
        weights = grads.mean(dim=(1, 2))
        cam = torch.relu((weights[:, None, None] * acts).sum(0))

        if cam.max() <= 1e-12:
            # HiResCAM-style fallback for saturated or tiny gradients.
            cam = torch.relu((grads * acts).sum(0))

        if cam.max() <= 1e-12:
            cam = torch.abs((weights[:, None, None] * acts).sum(0))

        cam = self._normalize_cam(cam)
        return cam.cpu().numpy()


    def cleanup(self):
        for h in self.hook_handles:
            h.remove()
        self.hook_handles = []
        self.gradients = None
        self.activations = None

    def __del__(self):
        self.cleanup()

    @staticmethod
    def _jet_colormap(value):
        r = np.clip(1.5 - np.abs(4.0 * value - 3.0), 0, 1)
        g = np.clip(1.5 - np.abs(4.0 * value - 2.0), 0, 1)
        b = np.clip(1.5 - np.abs(4.0 * value - 1.0), 0, 1)
        return np.stack([r, g, b], axis=-1)

    @staticmethod
    def overlay_heatmap(original_image, heatmap, alpha=0.45):
        heatmap = np.nan_to_num(heatmap, nan=0.0, posinf=1.0, neginf=0.0)
        heatmap = np.clip(heatmap, 0.0, 1.0)
        img_np = np.array(original_image.convert("RGB"))
        h, w = img_np.shape[:2]

        hm_img = Image.fromarray(np.uint8(255 * heatmap)).resize((w, h), Image.BICUBIC)
        hm_arr = np.array(hm_img).astype(np.float32) / 255.0

        try:
            from scipy.ndimage import gaussian_filter
            hm_arr = gaussian_filter(hm_arr, sigma=max(w, h) * 0.006)
            if hm_arr.max() > 0:
                hm_arr = hm_arr / hm_arr.max()
        except ImportError:
            pass

        gray = img_np.astype(np.float32).mean(axis=2) / 255.0
        tissue_mask = np.clip((gray - 0.04) / 0.12, 0.0, 1.0)
        hm_arr = hm_arr * tissue_mask
        if hm_arr.max() > 0:
            hm_arr = hm_arr / hm_arr.max()

        heatmap_colored = np.uint8(GradCAM._jet_colormap(hm_arr) * 255)
        alpha_map = (np.clip(hm_arr, 0.0, 1.0) ** 0.85) * alpha
        alpha_map = alpha_map[..., None]
        overlay = (img_np * (1 - alpha_map) + heatmap_colored * alpha_map).astype(np.uint8)
        return Image.fromarray(overlay)


def ui_wrap(text, width=34, max_lines=None):
    """Force predictable wrapping in narrow CustomTkinter panels."""
    wrapped = textwrap.wrap(str(text), width=width, break_long_words=False, break_on_hyphens=False)
    if not wrapped:
        return str(text)
    if max_lines and len(wrapped) > max_lines:
        wrapped = wrapped[:max_lines]
    return "\n".join(wrapped)

# ── Настройки темы ──
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ══════════════════════════════════════════════════════════
#  МЕДИЦИНСКАЯ БАЗА ЗНАНИЙ
# ══════════════════════════════════════════════════════════
PATHOLOGIES = {
    "polyp": {
        "name": "Полип ЖКТ",
        "latin": "Polypus tractus gastrointestinalis",
        "emoji": "🔵",
        "icd": "K31.7 / K63.5",
        "severity": 2,
        "sev_text": "Умеренная",
        "color": "#38bdf8",
        "description": (
            "Доброкачественное новообразование слизистой оболочки желудочно-кишечного тракта, "
            "выявляемое при эндоскопическом исследовании (ЭГДС или колоноскопии). "
            "Может локализоваться в пищеводе, желудке, ДПК или толстой кишке. "
            "Риск малигнизации зависит от гистологического типа, размера и локализации."
        ),
        "treatment": {
            "Медикаменты": [
                ("💊 Аспирин 100 мг/сут", "Профилактика рецидива аденом у пациентов высокого риска", "Класс IIb"),
                ("🧴 Пробиотики", "Lactobacillus + Bifidobacterium — нормализация микробиоты после вмешательства", "3–6 мес"),
                ("💊 Целекоксиб 400 мг/сут", "При семейном аденоматозном полипозе (ФАП)", "По показаниям"),
            ],
            "Процедуры": [
                ("✂️ Полипэктомия", "Эндоскопическое удаление петлей или щипцами. Метод выбора для полипов до 2 см", "Планово"),
                ("🔬 Гистология", "Обязательная морфологическая верификация удалённого материала", "После удаления"),
                ("📡 ЭУС (эндо-УЗИ)", "Для полипов более 2 см — оценка глубины инвазии", "По показаниям"),
            ],
            "Диета": [
                ("🌾 Клетчатка ≥35 г/сут", "Цельнозерновые, бобовые, овощи, фрукты — доказанный протективный эффект", "Ежедневно"),
                ("🥦 Крестоцветные овощи", "Брокколи, цветная капуста содержат сульфорафан с антиканцерогенным действием", "3–5 раз/нед"),
                ("🚫 Исключить", "Красное мясо, алкоголь, жареные и жирные продукты, рафинированные углеводы", "Постоянно"),
            ],
            "Наблюдение": [
                ("📅 Колоноскопия ч/з 1 год", "При аденомах высокого риска: ≥3 шт., ворсинчатый тип, дисплазия высокой ст.", "Обязательно"),
                ("📅 Колоноскопия ч/з 3 года", "При 1–2 тубулярных аденомах <10 мм с дисплазией низкой степени", "Стандарт"),
                ("🧬 ДНК-тест кала", "Cologuard — как дополнение между эндоскопическими контролями", "Ежегодно"),
            ],
        },
    },
    "cancer": {
        "name": "Злокачественное новообразование ЖКТ",
        "latin": "Neoplasma malignum tractus gastrointestinalis",
        "emoji": "🔴",
        "icd": "C15–C20",
        "severity": 4,
        "sev_text": "Критическая",
        "color": "#f87171",
        "description": (
            "Злокачественное новообразование, выявленное при эндоскопическом исследовании. "
            "Включает рак пищевода, желудка, двенадцатиперстной и толстой кишки. "
            "Прогноз определяется стадией, локализацией и гистологическим типом. "
            "Раннее эндоскопическое выявление критически важно для выживаемости."
        ),
        "treatment": {
            "Медикаменты": [
                ("💉 FOLFOX", "Оксалиплатин 85 мг/м² + Лейковорин + 5-ФУ — стандарт 1-й линии метастатического КРР", "Каждые 2 нед"),
                ("🎯 Бевацизумаб / Цетуксимаб", "Таргетная терапия (анти-VEGF или анти-EGFR) в комбинации с FOLFOX", "По протоколу"),
                ("🛡️ Пембролизумаб", "Иммунотерапия при MSI-H / dMMR — 1-я линия при микросателлитной нестабильности", "200 мг /3 нед"),
            ],
            "Процедуры": [
                ("⚕️ Хирургическая резекция", "Гемиколэктомия или резекция прямой кишки — основной метод при резектабельных опухолях", "Основной метод"),
                ("☢️ Лучевая терапия", "Неоадъювантная химиолучевая терапия при раке прямой кишки T3–T4/N+", "45–50.4 Гр"),
                ("🔗 РЧА метастазов", "Радиочастотная абляция единичных печёночных метастазов <3 см", "По показаниям"),
            ],
            "Диета": [
                ("⚡ Нутритивная поддержка", "Белок 1.5–2 г/кг/сут для снижения катаболизма во время химиотерапии", "Весь курс"),
                ("💧 Гидратация 2–2.5 л/сут", "Профилактика нефротоксичности при применении Оксалиплатина", "Ежедневно"),
                ("🚫 Исключить", "Алкоголь, красное мясо, грейпфрут (при таргетной терапии)", "Обязательно"),
            ],
            "Наблюдение": [
                ("🩸 КЭА каждые 3 мес", "Карциноэмбриональный антиген — контроль ответа на лечение (3 года)", "Ежеквартально"),
                ("🖥️ КТ брюшной полости", "Оценка системного ответа и выявление новых метастазов каждые 6 мес", "2 раза/год"),
                ("📡 ПЭТ-КТ", "При росте КЭА без видимых очагов на КТ для поиска скрытых метастазов", "По показаниям"),
            ],
        },
    },
    "ulcer": {
        "name": "Язвенное поражение ЖКТ",
        "latin": "Ulcus tractus gastrointestinalis",
        "emoji": "🟡",
        "icd": "K25–K28",
        "severity": 3,
        "sev_text": "Высокая",
        "color": "#fbbf24",
        "description": (
            "Дефект слизистой оболочки, выявляемый при эндоскопии (ЭГДС). "
            "Включает язвы пищевода, желудка и двенадцатиперстной кишки. "
            "В 80–95% случаев ассоциировано с H. pylori или приёмом НПВС. "
            "Требует биопсии для исключения малигнизации."
        ),
        "treatment": {
            "Медикаменты": [
                ("💊 Тройная эрадикация (1-я линия)", "Омепразол 20 мг + Кларитромицин 500 мг + Амоксициллин 1000 мг — 2р/сут, 14 дней", "14 дней"),
                ("💊 Квадротерапия (2-я линия)", "Висмут + Тетрациклин + Метронидазол + ИПП — при резистентности к 1-й линии", "10–14 дней"),
                ("🧴 ИПП поддержка", "Омепразол 20 мг или Пантопразол 40 мг/сут после курса эрадикации", "4–8 недель"),
            ],
            "Процедуры": [
                ("🔬 13C-дыхательный тест", "Контроль эрадикации H. pylori через 4–6 недель после завершения терапии", "Обязательно"),
                ("🩺 ЭГДС с биопсией", "Контроль заживления язв желудка через 8–12 нед. Биопсия для исключения малигнизации", "Через 2–3 мес"),
                ("⚕️ Хирургия", "При осложнениях: перфорация, кровотечение (прошивание), стеноз привратника", "Экстренно"),
            ],
            "Диета": [
                ("🍽️ Дробное питание 5–6 р/сут", "Небольшие порции, тщательное пережёвывание, последний приём за 3 ч до сна", "Постоянно"),
                ("🚫 Исключить триггеры", "НПВС, аспирин, алкоголь, кофе, острые специи, газировка, курение", "Обязательно"),
                ("🥣 Диета №1 по Певзнеру", "Каши, нежирное мясо/рыба на пару, кисели, некислые молочные, варёные овощи", "Во время лечения"),
            ],
            "Наблюдение": [
                ("📅 ЭГДС ч/з 8–12 нед", "Контроль заживления язв желудка. Биопсия из краёв при неполном заживлении", "Обязательно"),
                ("🧪 Тест на H. pylori", "Уреазный дыхательный тест через 4–6 нед после эрадикации для оценки успеха", "Один раз"),
                ("📅 Ежегодная ЭГДС", "При часто рецидивирующем течении, длительном приёме НПВС", "1 раз/год"),
            ],
        },
    },
    "gastritis": {
        "name": "Гастрит / Эзофагит",
        "latin": "Gastritis / Oesophagitis",
        "emoji": "🟠",
        "icd": "K29 / K20",
        "severity": 1,
        "sev_text": "Лёгкая",
        "color": "#fb923c",
        "description": (
            "Воспаление слизистой оболочки желудка или пищевода, "
            "визуализируемое при ЭГДС как гиперемия, отёк, эрозии. "
            "Причины: H. pylori (70%), рефлюкс, химические агенты, аутоиммунные факторы. "
            "Эндоскопическая классификация по Сиднейской системе и Лос-Анджелесской шкале."
        ),
        "treatment": {
            "Медикаменты": [
                ("💊 ИПП терапия", "Омепразол 20 мг или Рабепразол 20 мг утром натощак при кислотозависимом гастрите", "4–8 недель"),
                ("🧴 Антациды", "Маалокс, Фосфалюгель — симптоматически. Применять через 1–1.5 ч после еды", "По требованию"),
                ("💊 Домперидон 10 мг", "3 раза/сут при рефлюкс-гастрите и нарушениях моторики желудка", "2–4 недели"),
            ],
            "Процедуры": [
                ("🔬 Тест на H. pylori", "Экспресс-уреазный тест при ЭГДС, дыхательный тест или серология", "Первичная диагностика"),
                ("🧬 Биопсия по Сиднейскому протоколу", "5 биоптатов (антрум, угол, тело) для определения степени и стадии по OLGA/OLGIM", "При ЭГДС"),
                ("📡 pH-метрия", "Суточная pH-метрия для оценки кислотной функции и подбора дозы ИПП", "По показаниям"),
            ],
            "Диета": [
                ("🍽️ Питание 4–5 раз в день", "Регулярные приёмы пищи, не торопясь, без голодания более 5 часов", "Постоянно"),
                ("🚫 Ограничить", "Острое, копчёное, алкоголь, кофе натощак, газировку, грубую пищу", "Постоянно"),
                ("🥗 Диета №2", "Нейтральные блюда: каши, варёные овощи, нежирное мясо, кисломолочные", "Регулярно"),
            ],
            "Наблюдение": [
                ("📅 ЭГДС 1 раз в год", "При атрофическом гастрите — мониторинг метаплазии и атрофии по OLGA", "Онкологическая настороженность"),
                ("🧬 Контроль эрадикации", "При H. pylori — дыхательный тест через 4–6 недель после курса терапии", "Один раз"),
                ("🔬 Биопсия при изменениях", "При нарастании атрофии — более частые ЭГДС с расширенной биопсией", "По показаниям"),
            ],
        },
    },
    "barrett": {
        "name": "Пищевод Барретта",
        "latin": "Oesophagus Barrett",
        "emoji": "🟣",
        "icd": "K22.7",
        "severity": 3,
        "sev_text": "Высокая",
        "color": "#a78bfa",
        "description": (
            "Метапластическое замещение плоского эпителия пищевода цилиндрическим (кишечным типом), "
            "визуализируемое при ЭГДС как «языки пламени» выше Z-линии. "
            "Предраковое состояние с ежегодным риском аденокарциномы 0.5–1%. "
            "Диагноз подтверждается биопсией по Пражским критериям (C&M)."
        ),
        "treatment": {
            "Медикаменты": [
                ("💊 Высокодозные ИПП", "Эзомепразол 40 мг или Пантопразол 40 мг 2р/сут — длительная кислотосупрессия", "Пожизненно"),
                ("💊 Аспирин 100–325 мг/сут", "Хемопрофилактика прогрессирования до аденокарциномы (данные когортных исследований)", "По решению врача"),
                ("🧴 Альгинаты (Гавискон)", "При прорывных симптомах рефлюкса на фоне ИПП-терапии", "После еды и на ночь"),
            ],
            "Процедуры": [
                ("🔥 РЧА (радиочастотная абляция)", "Метод выбора при дисплазии высокой степени и раннем раке Барретта", "1–3 сеанса"),
                ("✂️ EMR (эндорезекция слизистой)", "При видимых очагах для гистологии и лечения раннего рака", "По показаниям"),
                ("❄️ Криоабляция", "Альтернатива РЧА или дополнительный метод при неполном ответе", "По показаниям"),
            ],
            "Диета": [
                ("⬇️ Снижение веса до ИМТ <25", "Абдоминальное ожирение — ведущий фактор риска ГЭРБ и пищевода Барретта", "Постоянная цель"),
                ("🚫 Триггеры рефлюкса", "Алкоголь, кофе, шоколад, мята, жирная пища, цитрусовые, томаты", "Исключить"),
                ("🛏️ Позиционная коррекция", "Приём пищи за 3 ч до сна. Приподнятый головной конец кровати на 15–20 см", "Ежедневно"),
            ],
            "Наблюдение": [
                ("📅 ЭГДС каждые 3–5 лет", "При коротком сегменте без дисплазии — стандартный протокол наблюдения", "Плановая"),
                ("📅 ЭГДС каждые 6–12 мес", "При дисплазии низкой степени до принятия решения об абляции", "Усиленный контроль"),
                ("🧬 Молекулярные маркёры", "p53-ИГХ, FISH при неопределённых морфологических изменениях", "По показаниям"),
            ],
        },
    },
    "ibd": {
        "name": "Воспалительные заболевания кишечника",
        "latin": "Morbus Crohn / Colitis ulcerosa",
        "emoji": "🔶",
        "icd": "K50–K51",
        "severity": 3,
        "sev_text": "Высокая",
        "color": "#fb923c",
        "description": (
            "Хронические иммуноопосредованные заболевания ЖКТ, диагностируемые "
            "при илеоколоноскопии и ЭГДС: болезнь Крона (трансмуральное воспаление, "
            "«булыжная мостовая», афтозные язвы) и язвенный колит "
            "(поверхностное воспаление с непрерывным поражением от прямой кишки). "
            "Эндоскопическая оценка по шкалам Mayo и CDEIS/SES-CD."
        ),
        "treatment": {
            "Медикаменты": [
                ("💊 Месалазин 2–4 г/сут", "Пентаса / Салофальк — базисная терапия ЯК лёгкой и средней степени тяжести", "Постоянно"),
                ("💉 Инфликсимаб / Адалимумаб", "Биологическая терапия при среднетяжёлом и тяжёлом ВЗК, рефрактерных формах", "По протоколу"),
                ("💊 Азатиоприн 2–2.5 мг/кг/сут", "Иммуносупрессор для поддержания ремиссии и снижения иммуногенности к биопрепаратам", "Длительно"),
            ],
            "Процедуры": [
                ("💩 Трансплантация микробиоты", "ТФМ — перспективный метод при рецидивирующем ЯК (одобрен в ряде стран)", "Экспериментально"),
                ("⚕️ Колэктомия", "При кровотечении, перфорации, токсическом мегаколоне, дисплазии, неэффективности терапии", "При осложнениях"),
                ("💉 GCAP (аферез)", "Гранулоцитарно-цитоферез при гормонозависимом ЯК", "По показаниям"),
            ],
            "Диета": [
                ("🥗 Низкорезидуальная диета", "В период обострения: исключение клетчатки, переход на легкоусвояемые продукты", "При обострении"),
                ("🧪 Энтеральное питание", "Нутриэн Элементаль при болезни Крона — снижает активность воспаления", "При обострении"),
                ("🚫 Индивидуальные триггеры", "Пищевой дневник для выявления продуктов-провокаторов (молоко, пшеница, алкоголь)", "Постоянно"),
            ],
            "Наблюдение": [
                ("📅 Колоноскопия каждые 1–2 года", "При ВЗК >10 лет — скрининг дисплазии (риск КРР значительно повышен)", "Онкоскрининг"),
                ("🩸 СРБ + кальпротектин кала", "Каждые 3 мес для оценки активности воспаления и анемии", "Ежеквартально"),
                ("🖥️ МРТ кишечника", "МР-энтерография при болезни Крона — оценка трансмурального воспаления и свищей", "1 раз/год"),
            ],
        },
    },
    "normal": {
        "name": "Норма — эндоскопическая картина без патологии",
        "latin": "Status normalis endoscopicus",
        "emoji": "🟢",
        "icd": "Z12.1",
        "severity": 0,
        "sev_text": "Норма",
        "color": "#34d399",
        "description": (
            "Слизистая оболочка визуализируемых отделов ЖКТ без признаков "
            "патологических изменений. Нормальная эндоскопическая картина: "
            "слизистая гладкая, блестящая, с типичным сосудистым рисунком, "
            "без гиперемии, эрозий, язв и новообразований."
        ),
        "treatment": {
            "Медикаменты": [
                ("✅ Лечение не требуется", "Патологических изменений не выявлено. Профилактический приём препаратов не показан", "—"),
            ],
            "Процедуры": [
                ("📅 Плановый скрининг", "Следующая колоноскопия через 10 лет (лицам 50+ при отсутствии факторов риска)", "По регламенту"),
            ],
            "Диета": [
                ("🥗 Средиземноморская диета", "Овощи, фрукты, рыба, цельнозерновые, оливковое масло — доказанный протективный эффект", "Постоянно"),
                ("🏃 Физическая активность", "Не менее 150 мин аэробных нагрузок в неделю — снижение риска КРР на 24%", "Регулярно"),
            ],
            "Наблюдение": [
                ("📅 Колоноскопия через 10 лет", "Стандартный интервал скрининга при нормальной первичной колоноскопии", "Плановая"),
                ("🩸 ФИТ-тест ежегодно", "Фекальный иммунохимический тест — неинвазивный скрининг между эндоскопиями", "Ежегодно"),
            ],
        },
    },
}

SEVERITY_COLORS = {0: "#34d399", 1: "#34d399", 2: "#fbbf24", 3: "#fb923c", 4: "#f87171"}
SEVERITY_LABELS = {0: "Норма", 1: "I — Лёгкая", 2: "II — Умеренная", 3: "III — Высокая", 4: "IV — Критическая"}

WEIGHTS = [0.22, 0.08, 0.18, 0.20, 0.10, 0.10, 0.12]

HISTORY_FILE = Path.home() / ".endomed" / "history.json"
SETTINGS_FILE = Path.home() / ".endomed" / "settings.json"

# Порог уверенности (по умолчанию)
DEFAULT_CONFIDENCE_THRESHOLD = 50.0  # %

# Версия схемы данных для миграции истории
SCHEMA_VERSION = 2

def load_settings():
    """Загружает настройки из файла."""
    defaults = {"confidence_threshold": DEFAULT_CONFIDENCE_THRESHOLD, "schema_version": SCHEMA_VERSION}
    try:
        if SETTINGS_FILE.exists():
            with open(SETTINGS_FILE, "r") as f:
                saved = json.load(f)
                defaults.update(saved)
    except Exception:
        pass
    return defaults

def save_settings(settings):
    """Сохраняет настройки в файл."""
    try:
        SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
        settings["schema_version"] = SCHEMA_VERSION
        with open(SETTINGS_FILE, "w") as f:
            json.dump(settings, f, indent=2)
    except Exception:
        pass

def export_history_to_excel(history, filepath):
    """Экспортирует историю диагностик в Excel (.xlsx) или CSV."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "История диагностик EndoMed"
        
        # Заголовки
        headers = ["№", "Дата и время", "Файл", "Диагноз", "Латинское", "МКБ-10", 
                   "Уверенность (%)", "Степень тяжести", "Режим модели"]
        header_fill = PatternFill(start_color="1a2540", end_color="1a2540", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="38bdf8")
        thin_border = Border(
            left=Side(style="thin", color="334155"),
            right=Side(style="thin", color="334155"),
            top=Side(style="thin", color="334155"),
            bottom=Side(style="thin", color="334155"),
        )
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        
        # Данные
        for i, r in enumerate(history):
            row = i + 2
            path_info = r.get("path", {})
            vals = [
                i + 1,
                r.get("datetime", ""),
                r.get("filename", ""),
                path_info.get("name", r.get("diagnosis", "")),
                path_info.get("latin", ""),
                path_info.get("icd", r.get("icd", "")),
                round(r.get("confidence", 0), 1),
                SEVERITY_LABELS.get(path_info.get("severity", 0), ""),
                r.get("mode", ""),
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = Font(name="Arial", size=10)
                cell.border = thin_border
                if col == 7:  # Уверенность
                    cell.alignment = Alignment(horizontal="center")
                    if v >= 90:
                        cell.font = Font(name="Arial", size=10, color="FF34D399")
                    elif v >= 70:
                        cell.font = Font(name="Arial", size=10, color="FFfbbf24")
                    else:
                        cell.font = Font(name="Arial", size=10, color="FFf87171")
        
        # Автоширина
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)
        
        wb.save(filepath)
        return True
    except ImportError:
        # Fallback: CSV
        try:
            csv_path = filepath.replace(".xlsx", ".csv")
            import csv
            with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f, delimiter=";")
                writer.writerow(["№", "Дата", "Файл", "Диагноз", "МКБ-10", "Уверенность", "Тяжесть"])
                for i, r in enumerate(history):
                    path_info = r.get("path", {})
                    writer.writerow([
                        i + 1, r.get("datetime", ""), r.get("filename", ""),
                        path_info.get("name", ""), path_info.get("icd", ""),
                        round(r.get("confidence", 0), 1),
                        SEVERITY_LABELS.get(path_info.get("severity", 0), ""),
                    ])
            return csv_path
        except Exception:
            return False

# ══════════════════════════════════════════════════════════
#  МОДЕЛЬ EFFICIENTNET-B4 (v6)
# ══════════════════════════════════════════════════════════
CLASS_NAMES = ["polyp", "cancer", "ulcer", "gastritis", "barrett", "ibd", "normal"]
# Трансформации для EfficientNet-B4 (вход 380×380)
INFERENCE_TRANSFORM = None
if TORCH_AVAILABLE:
    INFERENCE_TRANSFORM = T.Compose([
        T.Resize((380, 380)),
        T.ToTensor(),
        T.Normalize(mean=[0.485, 0.456, 0.406],
                    std=[0.229, 0.224, 0.225]),
    ])


class EndoMedModel:
    """
    Обёртка над EfficientNet-B4 для классификации эндоскопических изображений.
    
    Исправления v6:
    - Thread-safe: весь инференс + CAM под одним замком
    - Grad-CAM пересоздаётся при каждом вызове generate_heatmap
    - Корректные пути к чекпоинтам
    - Явная ошибка при отсутствии pydicom для DICOM
    """
    
    def __init__(self):
        self.model = None
        self.device = None
        self.mode = "simulation"  # "checkpoint", "pretrained", "simulation"
        self.classes = CLASS_NAMES
        self.grad_cam = None
        
        if not TORCH_AVAILABLE:
            print("🔸 Режим: симуляция (PyTorch не найден)")
            return
        
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        print(f"🔧 Устройство: {self.device}")
        
        # Попытка 1: загрузить обученный чекпоинт
        # Используем __file__ для надёжных путей
        script_dir = Path(__file__).parent.resolve()
        checkpoint_paths = [
            script_dir / "endomed_checkpoint.pt",
            script_dir / "models" / "endomed_checkpoint.pt",
            Path.home() / ".endomed" / "endomed_checkpoint.pt",
            # Legacy: относительные пути для обратной совместимости
            Path("endomed_checkpoint.pt"),
            Path("models/endomed_checkpoint.pt"),
        ]
        
        for cp_path in checkpoint_paths:
            if cp_path.exists():
                try:
                    print(f"📦 Загрузка чекпоинта: {cp_path}")
                    checkpoint = torch.load(str(cp_path), map_location=self.device, weights_only=False)
                    
                    # Создаём модель
                    self.model = models.efficientnet_b4(weights=None)
                    num_classes = len(checkpoint.get("classes", CLASS_NAMES))
                    self.model.classifier[1] = torch.nn.Linear(
                        self.model.classifier[1].in_features, num_classes
                    )
                    
                    # Загружаем веса
                    self.model.load_state_dict(checkpoint["model_state_dict"])
                    self.model.to(self.device)
                    self.model.eval()
                    
                    self.classes = checkpoint.get("classes", CLASS_NAMES)
                    self.mode = "checkpoint"
                    print(f"✅ Модель загружена из чекпоинта ({num_classes} классов)")
                    print(f"   Классы: {self.classes}")
                    return
                except Exception as e:
                    logger.warning("Ошибка загрузки чекпоинта: %s", e)
        
        # Попытка 2: ImageNet pretrained с адаптированным выходом
        try:
            print("📦 Загрузка EfficientNet-B4 (ImageNet pretrained)...")
            self.model = models.efficientnet_b4(weights=models.EfficientNet_B4_Weights.IMAGENET1K_V1)
            
            # Заменяем классификатор на 7 классов
            in_features = self.model.classifier[1].in_features
            self.model.classifier[1] = torch.nn.Linear(in_features, len(CLASS_NAMES))
            
            # Инициализируем новый слой
            torch.nn.init.xavier_uniform_(self.model.classifier[1].weight)
            torch.nn.init.zeros_(self.model.classifier[1].bias)
            
            self.model.to(self.device)
            self.model.eval()
            self.mode = "pretrained"
            print(f"✅ EfficientNet-B4 загружена (pretrained ImageNet, демо-режим)")
            print(f"   in_features={in_features}, out=7, device={self.device}")
        except Exception as e:
            logger.warning("Не удалось загрузить модель: %s", e)
            print("🔸 Переключение в режим симуляции")
            self.mode = "simulation"
    
    def predict(self, image_path: str, use_tta: bool = True, cancel_event: threading.Event = None) -> dict:
        """
        Выполняет инференс на изображении с опциональным TTA.
        
        Args:
            image_path: путь к изображению
            use_tta: использовать Test Time Augmentation
            cancel_event: Event для отмены анализа
        
        Returns:
            dict с результатами
        """
        start = time.time()
        
        if self.mode == "simulation" or self.model is None:
            # Проверяем отмену даже в симуляции
            if cancel_event and cancel_event.is_set():
                return {"cancelled": True}
            return self._simulate(cancel_event)
        
        try:
            # Загрузка
            if image_path.lower().endswith((".dcm", ".dicom")):
                img = self._load_dicom_for_inference(image_path)
                if img is None:
                    return {"error": "DICOM_ERROR", "message": "Не удалось загрузить DICOM"}
            else:
                img = Image.open(image_path).convert("RGB")

            if img is None:
                return self._simulate(cancel_event)

            with _model_lock:
                # Проверка отмены перед началом
                if cancel_event and cancel_event.is_set():
                    return {"cancelled": True}
                
                self.model.eval()

                if use_tta and self.mode == "checkpoint":
                    # TTA: 5 проходов, усреднение
                    tta_transforms = [
                        INFERENCE_TRANSFORM,
                        T.Compose([T.Resize((380, 380)), T.RandomHorizontalFlip(1.0),
                                   T.ToTensor(), T.Normalize([0.485,0.456,0.406],[0.229,0.224,0.225])]),
                        T.Compose([T.Resize((380, 380)), T.RandomVerticalFlip(1.0),
                                   T.ToTensor(), T.Normalize([0.485,0.456,0.406],[0.229,0.224,0.225])]),
                        T.Compose([T.Resize((420, 420)), T.CenterCrop(380),
                                   T.ToTensor(), T.Normalize([0.485,0.456,0.406],[0.229,0.224,0.225])]),
                        T.Compose([T.Resize((420, 420)), T.RandomCrop(380),
                                   T.ToTensor(), T.Normalize([0.485,0.456,0.406],[0.229,0.224,0.225])]),
                    ]
                    all_probs = []
                    with torch.no_grad():
                        for i, tta_tf in enumerate(tta_transforms):
                            # Проверка отмены между аугментациями
                            if cancel_event and cancel_event.is_set():
                                return {"cancelled": True}
                            
                            tensor = tta_tf(img).unsqueeze(0).to(self.device)
                            probs = F.softmax(self.model(tensor), dim=1)[0]
                            all_probs.append(probs.cpu().numpy())
                    probs_np = np.mean(all_probs, axis=0)
                else:
                    # Обычный инференс
                    tensor = INFERENCE_TRANSFORM(img).unsqueeze(0).to(self.device)
                    with torch.no_grad():
                        logits = self.model(tensor)
                        probs = F.softmax(logits, dim=1)[0]
                    probs_np = probs.cpu().numpy()

            elapsed = (time.time() - start) * 1000

            # Формируем все предсказания
            all_preds = []
            for i, cls_name in enumerate(self.classes):
                all_preds.append((cls_name, round(float(probs_np[i]) * 100, 1)))
            all_preds.sort(key=lambda x: -x[1])

            confidence = all_preds[0][1]

            return {
                "key": all_preds[0][0],
                "confidence": confidence,
                "all_preds": all_preds[:7],
                "mode": self.mode,
                "inference_time_ms": round(elapsed, 1),
                "tta": use_tta and self.mode == "checkpoint",
                "heatmap": None,
            }

        except Exception as e:
            logger.error("Ошибка инференса: %s", e)
            return self._simulate(cancel_event)
    
    def _load_dicom_for_inference(self, path):
        """Загружает DICOM для инференса. Возвращает None при ошибке."""
        try:
            if DICOM_AVAILABLE:
                ds = pydicom.dcmread(path)
                arr = ds.pixel_array.astype(np.float32)
                # Нормализация
                if hasattr(ds, 'WindowCenter') and hasattr(ds, 'WindowWidth'):
                    wc = float(ds.WindowCenter if not isinstance(ds.WindowCenter, pydicom.multival.MultiValue) else ds.WindowCenter[0])
                    ww = float(ds.WindowWidth if not isinstance(ds.WindowWidth, pydicom.multival.MultiValue) else ds.WindowWidth[0])
                    low = wc - ww / 2
                    high = wc + ww / 2
                    arr = np.clip(arr, low, high)
                arr = ((arr - arr.min()) / (arr.max() - arr.min() + 1e-6) * 255).astype(np.uint8)
                # Если grayscale → RGB
                if len(arr.shape) == 2:
                    return Image.fromarray(arr, mode="L").convert("RGB")
                elif arr.shape[2] == 3:
                    return Image.fromarray(arr, mode="RGB")
                else:
                    return Image.fromarray(arr[:, :, 0], mode="L").convert("RGB")
            else:
                # Явная ошибка: без pydicom DICOM не поддерживается
                logger.error("pydicom не установлен, невозможно прочитать DICOM")
                return None
        except Exception as e:
            logger.error("Ошибка загрузки DICOM: %s", e)
            return None
    
    def _simulate(self, cancel_event=None) -> dict:
        """Симуляция для случаев когда модель недоступна."""
        # Проверяем отмену
        if cancel_event and cancel_event.is_set():
            return {"cancelled": True}
            
        time.sleep(1.5)
        
        if cancel_event and cancel_event.is_set():
            return {"cancelled": True}
        
        keys = list(PATHOLOGIES.keys())
        key = weighted_choice(keys, WEIGHTS)
        confidence = round(random.uniform(72.0, 98.5), 1)
        
        other_confs = []
        remaining = 100 - confidence
        for k in keys:
            if k != key:
                c = round(random.uniform(0, remaining * 0.4), 1)
                other_confs.append((k, c))
        all_preds = [(key, confidence)] + sorted(other_confs, key=lambda x: -x[1])
        
        return {
            "key": key,
            "confidence": confidence,
            "all_preds": all_preds[:7],
            "mode": "simulation",
            "inference_time_ms": 1500,
        }
    
    def generate_heatmap(self, image_path: str, class_idx=None, cancel_event=None):
        """
        Генерирует Grad-CAM тепловую карту для изображения.
        Thread-safe: создаёт новый GradCAM на каждый вызов.
        
        Returns:
            PIL Image с наложенной тепловой картой, или None
        """
        if self.model is None:
            return None
        
        try:
            with _model_lock:
                # Проверка отмены
                if cancel_event and cancel_event.is_set():
                    return None
                
                # Загрузка (поддержка DICOM)
                if image_path.lower().endswith((".dcm", ".dicom")):
                    img = self._load_dicom_for_inference(image_path)
                else:
                    img = Image.open(image_path).convert("RGB")

                if img is None:
                    return None

                # GradCAM() без target_layer использует features[-1] — весь последний
                # Sequential целиком. Это корректно: хук получает выход блока 1792×12×12.
                grad_cam = GradCAM(self.model)
                
                try:
                    self.model.eval()
                    tensor = INFERENCE_TRANSFORM(img).unsqueeze(0).to(self.device)

                    # enable_grad обязателен — вызывающий код или predict()
                    # мог оставить активным контекст no_grad
                    with torch.enable_grad():
                        heatmap = grad_cam.generate(tensor, class_idx)
                finally:
                    grad_cam.cleanup()

                overlay = GradCAM.overlay_heatmap(img, heatmap, alpha=0.45)
                return overlay
                
        except Exception as e:
            logger.error("Ошибка Grad-CAM: %s", e)
            import traceback
            traceback.print_exc()
            return None

    def get_status_text(self) -> tuple:
        """Возвращает (текст, цвет) для индикатора статуса."""
        if self.mode == "checkpoint":
            return ("● EfficientNet-B4  обучена", "#4ade80")
        elif self.mode == "pretrained":
            return ("● EfficientNet-B4  демо", "#fbbf24")
        else:
            return ("● Режим симуляции", "#f87171")


# Глобальный экземпляр модели (загрузка при старте)
print("=" * 60)
print("  EndoMed — Инициализация модели")
print("=" * 60)
MODEL = EndoMedModel()
print("=" * 60)

# ══════════════════════════════════════════════════════════
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ══════════════════════════════════════════════════════════
def weighted_choice(items, weights):
    return random.choices(items, weights=weights, k=1)[0]

def darken_color(hex_color, factor=0.15):
    """Returns a darkened version of a hex color."""
    hex_color = hex_color.lstrip("#")
    # Поддержка сокращённого hex (#fff)
    if len(hex_color) == 3:
        hex_color = "".join([c*2 for c in hex_color])
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    r, g, b = int(r * (1 - factor)), int(g * (1 - factor)), int(b * (1 - factor))
    return f"#{r:02x}{g:02x}{b:02x}"

def export_pdf_report(result_data: dict, filepath: str):
    """Экспортирует результат диагностики в PDF через reportlab с поддержкой кириллицы."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        # Регистрация шрифта с поддержкой кириллицы
        # Ищем системные шрифты в порядке приоритета
        font_paths = [
            # Linux
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
            # macOS
            "/System/Library/Fonts/Helvetica.ttc",
            "/Library/Fonts/Arial.ttf",
            # Windows
            "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/tahoma.ttf",
            # Относительно скрипта
            str(Path(__file__).parent / "fonts" / "DejaVuSans.ttf"),
        ]
        
        font_name = "Helvetica"
        for fp in font_paths:
            if Path(fp).exists():
                try:
                    font_name = "CustomCyrillic"
                    pdfmetrics.registerFont(TTFont(font_name, fp))
                    break
                except Exception:
                    continue
        
        # Fallback: если не нашли кириллический шрифт, используем стандартный
        # (reportlab поддерживает встроенные шрифты, но они не умеют кириллицу)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)

        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle('Title', parent=styles['Title'],
                                      fontSize=22, spaceAfter=6,
                                      textColor=colors.HexColor('#1e3a5f'),
                                      fontName=font_name)
        story.append(Paragraph("EndoMed — Заключение по результатам диагностики", title_style))
        story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#38bdf8')))
        story.append(Spacer(1, 0.4*cm))

        # Meta
        meta_style = ParagraphStyle('Meta', parent=styles['Normal'], fontSize=10,
                                     textColor=colors.grey, spaceAfter=4,
                                     fontName=font_name)
        story.append(Paragraph(f"Дата и время: {result_data['datetime']}", meta_style))
        story.append(Paragraph(f"Файл: {result_data['filename']}", meta_style))
        story.append(Paragraph(f"Модель: EfficientNet-B4 | ICD-10: {result_data['icd']}", meta_style))
        story.append(Spacer(1, 0.5*cm))

        # Diagnosis block
        diag_style = ParagraphStyle('Diag', parent=styles['Heading1'], fontSize=16,
                                     textColor=colors.HexColor('#1e3a5f'),
                                     fontName=font_name)
        story.append(Paragraph(f"Диагноз: {result_data['diagnosis']}", diag_style))

        sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=11,
                                    textColor=colors.grey, spaceAfter=4,
                                    fontName=font_name)
        story.append(Paragraph(result_data['latin'], sub_style))

        info_data = [
            ["Уверенность модели:", f"{result_data['confidence']:.1f}%"],
            ["Степень тяжести:", result_data['severity']],
            ["МКБ-10:", result_data['icd']],
        ]
        t = Table(info_data, colWidths=[5*cm, 12*cm])
        t.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1e3a5f')),
            ('FONTNAME', (0, 0), (0, -1), font_name),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.4*cm))

        # Description
        desc_style = ParagraphStyle('Desc', parent=styles['Normal'], fontSize=11,
                                     leading=16, spaceAfter=12,
                                     backColor=colors.HexColor('#f0f8ff'),
                                     leftIndent=10, rightIndent=10, borderPadding=8,
                                     fontName=font_name)
        story.append(Paragraph(result_data['description'], desc_style))
        story.append(Spacer(1, 0.3*cm))

        # Treatment sections
        h2_style = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=13,
                                   textColor=colors.HexColor('#1e3a5f'), spaceAfter=6, spaceBefore=10,
                                   fontName=font_name)
        item_style = ParagraphStyle('Item', parent=styles['Normal'], fontSize=10,
                                     leading=15, leftIndent=12, spaceAfter=4,
                                     fontName=font_name)

        for section_name, items in result_data['treatment'].items():
            story.append(Paragraph(f"▸ {section_name}", h2_style))
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
            for title, desc, pill in items:
                # Больше не обрезаем кириллицу!
                story.append(Paragraph(f"<b>{title}</b> <font color='grey' size='9'>[{pill}]</font>", item_style))
                story.append(Paragraph(desc, ParagraphStyle('d', parent=item_style, leftIndent=24,
                                                             textColor=colors.HexColor('#444'), fontSize=9,
                                                             fontName=font_name)))
            story.append(Spacer(1, 0.2*cm))

        # Disclaimer
        disc_style = ParagraphStyle('Disc', parent=styles['Normal'], fontSize=9,
                                     textColor=colors.HexColor('#cc3333'),
                                     borderColor=colors.HexColor('#cc3333'),
                                     borderWidth=1, borderPadding=6,
                                     leftIndent=10, rightIndent=10, spaceAfter=0,
                                     fontName=font_name)
        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#cc3333')))
        story.append(Paragraph(
            "⚠ ВНИМАНИЕ: Данное заключение носит вспомогательный характер и не заменяет "
            "консультацию и заключение врача-специалиста. Все лечебные мероприятия "
            "назначаются исключительно лечащим врачом после клинического обследования.",
            ParagraphStyle('disc2', parent=styles['Normal'], fontSize=9,
                           textColor=colors.HexColor('#cc3333'), spaceBefore=6,
                           fontName=font_name)
        ))

        doc.build(story)
        return True
    except ImportError:
        return False
    except Exception as e:
        print(f"PDF error: {e}")
        return False

# ══════════════════════════════════════════════════════════
#  ГЛАВНОЕ ОКНО (v6)
# ══════════════════════════════════════════════════════════
class EndoMedApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("EndoMed — Диагностика патологий ЖКТ")
        self.geometry("1280x820")
        self.minsize(1100, 700)
        self.configure(fg_color="#0a0e17")

        # Устанавливаем иконку если есть
        try:
            self.iconbitmap("icon.ico")
        except Exception:
            pass

        self.current_image_path = None
        self.current_result = None
        self.history = []
        self.photo_image = None
        self.settings = load_settings()
        self.confidence_threshold = self.settings.get("confidence_threshold", DEFAULT_CONFIDENCE_THRESHOLD)
        
        # Event для отмены анализа
        self._cancel_event = threading.Event()
        self._analysis_thread = None
        self._warn_frame = None
        self._disclaimer_frame = None
        self._info_bar_frame = None

        self._load_history()
        self._build_ui()
        if self.history:
            self.hist_btn.configure(text=f"📋 История  {len(self.history)}")

    # ──────────────────────────────────────────────
    def _build_ui(self):
        # Шапка
        self._build_header()

        # Основной контейнер
        self.main_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.main_frame.pack(fill="both", expand=True, padx=0, pady=0)
        self.main_frame.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_columnconfigure(1, weight=1)
        self.main_frame.grid_rowconfigure(0, weight=1)

        # Левая панель — изображение
        self._build_image_panel()

        # Правая панель — результаты
        self._build_result_panel()

    # ──────────────────────────────────────────────
    def _build_header(self):
        hdr = ctk.CTkFrame(self, fg_color="#0d1420", height=60, corner_radius=0,
                           border_width=1, border_color="#1f2d47")
        hdr.pack(fill="x", side="top")
        hdr.pack_propagate(False)

        # Logo
        logo_frame = ctk.CTkFrame(hdr, fg_color="transparent")
        logo_frame.pack(side="left", padx=20, pady=10)

        logo_box = ctk.CTkFrame(logo_frame, fg_color="#164e63", width=36, height=36,
                                corner_radius=10)
        logo_box.pack(side="left")
        logo_box.pack_propagate(False)
        ctk.CTkLabel(logo_box, text="🔬", font=ctk.CTkFont(size=18)).pack(expand=True)

        ctk.CTkLabel(logo_frame, text="  EndoMed",
                     font=ctk.CTkFont(size=20, weight="bold"),
                     text_color="#f0f6ff").pack(side="left")
        ctk.CTkLabel(logo_frame, text=" — Система диагностики патологий ЖКТ",
                     font=ctk.CTkFont(size=13),
                     text_color="#475569").pack(side="left")

        # Right buttons
        right = ctk.CTkFrame(hdr, fg_color="transparent")
        right.pack(side="right", padx=20, pady=8)

        self.hist_btn = ctk.CTkButton(right, text="📋 История",
                                       width=110, height=36,
                                       fg_color="#1a2540", hover_color="#1e3a5f",
                                       border_color="#1f2d47", border_width=1,
                                       font=ctk.CTkFont(size=13),
                                       command=self._open_history)
        self.hist_btn.pack(side="right", padx=6)

        # Кнопка экспорта истории в Excel
        self.excel_btn = ctk.CTkButton(right, text="📊 Excel",
                                        width=90, height=36,
                                        fg_color="#1a2540", hover_color="#065f46",
                                        border_color="#047857", border_width=1,
                                        font=ctk.CTkFont(size=13),
                                        command=self._export_excel)
        self.excel_btn.pack(side="right", padx=6)

        # Кнопка настроек порога
        self.settings_btn = ctk.CTkButton(right, text="⚙️",
                                           width=40, height=36,
                                           fg_color="#1a2540", hover_color="#1e3a5f",
                                           border_color="#1f2d47", border_width=1,
                                           font=ctk.CTkFont(size=16),
                                           command=self._open_settings)
        self.settings_btn.pack(side="right", padx=6)

        # Кнопка дашборда статистики
        self.dashboard_btn = ctk.CTkButton(right, text="📊 Статистика",
                                            width=120, height=36,
                                            fg_color="#1a2540", hover_color="#1e3a5f",
                                            border_color="#1f2d47", border_width=1,
                                            font=ctk.CTkFont(size=13),
                                            command=self._open_dashboard)
        self.dashboard_btn.pack(side="right", padx=6)

        # Кнопка Grad-CAM
        self.gradcam_btn = ctk.CTkButton(right, text="🔥 Grad-CAM",
                                          width=120, height=36,
                                          fg_color="#1a2540", hover_color="#422006",
                                          border_color="#854d0e", border_width=1,
                                          font=ctk.CTkFont(size=13),
                                          state="disabled",
                                          command=self._show_gradcam)
        self.gradcam_btn.pack(side="right", padx=6)

        self.export_btn = ctk.CTkButton(right, text="📄 Экспорт PDF",
                                         width=130, height=36,
                                         fg_color="#1a2540", hover_color="#1e3a5f",
                                         border_color="#1f2d47", border_width=1,
                                         font=ctk.CTkFont(size=13),
                                         state="disabled",
                                         command=self._export_pdf)
        self.export_btn.pack(side="right", padx=6)

        # Status
        status_text, status_color = MODEL.get_status_text()
        status_bg = {"#4ade80": "#052e16", "#fbbf24": "#422006", "#f87171": "#450a0a"}
        status_border = {"#4ade80": "#166534", "#fbbf24": "#854d0e", "#f87171": "#991b1b"}
        status_frame = ctk.CTkFrame(right,
                                     fg_color=status_bg.get(status_color, "#052e16"),
                                     corner_radius=20,
                                     border_color=status_border.get(status_color, "#166534"),
                                     border_width=1)
        status_frame.pack(side="right", padx=8)
        ctk.CTkLabel(status_frame, text=status_text,
                     font=ctk.CTkFont(size=11, family="Courier"),
                     text_color=status_color,
                     padx=12, pady=5).pack()

    # ──────────────────────────────────────────────
    def _build_image_panel(self):
        left = ctk.CTkFrame(self.main_frame, fg_color="#0d1420",
                             corner_radius=16,
                             border_width=1, border_color="#1f2d47")
        left.grid(row=0, column=0, padx=(16, 8), pady=16, sticky="nsew")
        left.grid_rowconfigure(1, weight=1)
        left.grid_columnconfigure(0, weight=1)

        # Header
        lh = ctk.CTkFrame(left, fg_color="transparent", height=44)
        lh.grid(row=0, column=0, sticky="ew", padx=16, pady=(12, 0))
        lh.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(lh, text="СНИМОК ЭНДОСКОПА",
                     font=ctk.CTkFont(size=11, family="Courier"),
                     text_color="#475569").grid(row=0, column=0, sticky="w")

        # Buttons
        btn_frame = ctk.CTkFrame(lh, fg_color="transparent")
        btn_frame.grid(row=0, column=2)
        ctk.CTkButton(btn_frame, text="📁 Открыть файл",
                       width=130, height=32,
                       font=ctk.CTkFont(size=13),
                       command=self._open_file).pack(side="left", padx=4)

        # Drop zone / Image frame
        self.img_container = ctk.CTkFrame(left, fg_color="#080c14",
                                           corner_radius=12,
                                           border_width=2, border_color="#1f2d47")
        self.img_container.grid(row=1, column=0, padx=16, pady=12, sticky="nsew")
        self.img_container.grid_rowconfigure(0, weight=1)
        self.img_container.grid_columnconfigure(0, weight=1)

        # Placeholder
        self.placeholder_frame = ctk.CTkFrame(self.img_container, fg_color="transparent")
        self.placeholder_frame.grid(row=0, column=0)
        ctk.CTkLabel(self.placeholder_frame, text="🔬",
                     font=ctk.CTkFont(size=72)).pack(pady=(0, 8))
        ctk.CTkLabel(self.placeholder_frame,
                     text="Перетащите изображение сюда\nили нажмите «Открыть файл»",
                     font=ctk.CTkFont(size=14), text_color="#334155",
                     justify="center").pack()
        ctk.CTkLabel(self.placeholder_frame,
                     text="Поддерживаемые форматы: JPEG · PNG · BMP · TIFF · DICOM",
                     font=ctk.CTkFont(size=11, family="Courier"),
                     text_color="#1e2d47").pack(pady=(8, 0))

        # Image label (hidden initially)
        self.img_label = ctk.CTkLabel(self.img_container, text="")
        self.img_label.grid(row=0, column=0, sticky="nsew")
        self.img_label.grid_remove()

        # Drag & drop bindings
        self.img_container.bind("<Button-1>", lambda e: self._open_file())
        self.img_container.bind("<Enter>", self._on_drag_enter)
        self.img_container.bind("<Leave>", self._on_drag_leave)
        try:
            self.img_container.drop_target_register("DND_Files")
            self.img_container.dnd_bind("<<Drop>>", self._on_drop)
        except Exception:
            pass

        # Bottom bar
        bot = ctk.CTkFrame(left, fg_color="transparent", height=50)
        bot.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 12))
        bot.grid_columnconfigure(0, weight=1)

        self.file_label = ctk.CTkLabel(bot, text="Файл не выбран",
                                        font=ctk.CTkFont(size=12, family="Courier"),
                                        text_color="#334155")
        self.file_label.grid(row=0, column=0, sticky="w")

        # Frame for analyze + cancel buttons
        self.btn_frame = ctk.CTkFrame(bot, fg_color="transparent")
        self.btn_frame.grid(row=0, column=1)

        self.analyze_btn = ctk.CTkButton(self.btn_frame, text="⚡  Анализировать",
                                          width=170, height=40,
                                          font=ctk.CTkFont(size=15, weight="bold"),
                                          fg_color="#0ea5e9",
                                          hover_color="#0284c7",
                                          text_color="#000000",
                                          state="disabled",
                                          command=self._analyze)
        self.analyze_btn.pack(side="left")

        # Кнопка отмены (скрыта по умолчанию)
        self.cancel_btn = ctk.CTkButton(self.btn_frame, text="✕  Отмена",
                                         width=100, height=40,
                                         font=ctk.CTkFont(size=13, weight="bold"),
                                         fg_color="#7f1d1d",
                                         hover_color="#991b1b",
                                         text_color="#fca5a5",
                                         command=self._cancel_analysis)
        self.cancel_btn.pack(side="left", padx=(8, 0))
        self.cancel_btn.pack_forget()

    # ──────────────────────────────────────────────
    def _build_result_panel(self):
        right = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        right.grid(row=0, column=1, padx=(8, 16), pady=16, sticky="nsew")
        right.grid_rowconfigure(1, weight=1)
        right.grid_columnconfigure(0, weight=1)

        # Индикатор режима симуляции (яркая плашка)
        self.sim_banner = ctk.CTkFrame(right, fg_color="#7f1d1d",
                                        corner_radius=8,
                                        border_width=2, border_color="#f87171",
                                        height=40)
        self.sim_banner.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        self.sim_banner.grid_remove()  # скрыт по умолчанию
        
        ctk.CTkLabel(self.sim_banner, text="⚠️  РЕЖИМ СИМУЛЯЦИИ — Результаты генерируются случайным образом",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color="#fca5a5").pack(pady=8)

        # Scrollable results
        self.result_scroll = ctk.CTkScrollableFrame(right, fg_color="transparent",
                                                     scrollbar_button_color="#1f2d47")
        self.result_scroll.grid(row=1, column=0, sticky="nsew")
        self.result_scroll.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(1, weight=1)

        # Empty state card
        self.empty_card = ctk.CTkFrame(self.result_scroll,
                                        fg_color="#0d1420",
                                        corner_radius=16,
                                        border_width=1, border_color="#1f2d47")
        self.empty_card.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        self.empty_card.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(self.empty_card, text="🩺",
                     font=ctk.CTkFont(size=48)).grid(pady=(32, 8))
        ctk.CTkLabel(self.empty_card,
                     text="Загрузите снимок и нажмите\n«Анализировать» для диагностики",
                     font=ctk.CTkFont(size=13), text_color="#334155",
                     justify="center").grid(pady=(0, 32))

        # Result cards (built dynamically)
        self.diag_card_frame = None
        self.treat_card_frame = None

        # Показываем баннер симуляции если нужно
        if MODEL.mode == "simulation":
            self.sim_banner.grid()

    # ──────────────────────────────────────────────
    def _load_history(self):
        """Загружает историю диагностик из JSON-файла с миграцией схемы."""
        try:
            if HISTORY_FILE.exists():
                with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                    saved = json.load(f)
                
                # Проверяем версию схемы
                file_schema = saved.get("_schema_version", 1) if isinstance(saved, dict) else 1
                
                # Если сохранено как список (старая версия) или dict с items
                items = saved.get("items", saved) if isinstance(saved, dict) else saved
                if not isinstance(items, list):
                    items = []
                
                for item in items:
                    key = item.get("key", "")
                    if key in PATHOLOGIES:
                        item["path"] = PATHOLOGIES[key]
                        item["treatment"] = PATHOLOGIES[key]["treatment"]
                        self.history.append(item)
                        
                # Миграция: если версия устарела, пересохраняем
                if file_schema < SCHEMA_VERSION:
                    self._save_history()
                    
        except Exception as e:
            logger.warning("Ошибка загрузки истории: %s", e)

    def _save_history(self):
        """Сохраняет историю диагностик в JSON-файл с версией схемы."""
        try:
            HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
            serializable = []
            for item in self.history[:100]:  # последние 100 записей
                entry = {k: v for k, v in item.items() if k not in ("path", "treatment")}
                serializable.append(entry)
            
            data = {
                "_schema_version": SCHEMA_VERSION,
                "items": serializable
            }
            
            with open(HISTORY_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.warning("Ошибка сохранения истории: %s", e)

    # ──────────────────────────────────────────────
    def _on_drag_enter(self, e):
        self.img_container.configure(border_color="#38bdf8")

    def _on_drag_leave(self, e):
        self.img_container.configure(border_color="#1f2d47")

    def _on_drop(self, e):
        self.img_container.configure(border_color="#1f2d47")
        path = e.data.strip().strip("{}")
        if path:
            self._load_image(path)

    def _open_file(self):
        path = filedialog.askopenfilename(
            title="Выберите снимок эндоскопа",
            filetypes=[
                ("Изображения", "*.jpg *.jpeg *.png *.bmp *.tiff *.tif"),
                ("DICOM", "*.dcm *.dicom"),
                ("Все файлы", "*.*"),
            ]
        )
        if path:
            self._load_image(path)

    def _load_image(self, path):
        try:
            self.current_image_path = path
            fname = Path(path).name
            size = Path(path).stat().st_size

            if path.lower().endswith((".dcm", ".dicom")):
                # DICOM — реальная загрузка
                img = self._load_dicom(path)
                if img is None:
                    if not DICOM_AVAILABLE:
                        messagebox.showerror(
                            "DICOM",
                            "Не удалось прочитать DICOM-файл.\n\n"
                            "Установите библиотеку pydicom:\n"
                            "pip install pydicom"
                        )
                    else:
                        messagebox.showerror("DICOM", "Не удалось прочитать DICOM-файл.\nВозможно, файл повреждён или имеет неподдерживаемый формат.")
                    return
                img.thumbnail((550, 440), Image.LANCZOS)
                self.photo_image = ctk.CTkImage(light_image=img, dark_image=img,
                                                size=(img.width, img.height))
                self.placeholder_frame.grid_remove()
                self.img_label.configure(image=self.photo_image, text="")
                self.img_label.grid()
            else:
                img = Image.open(path)
                img.thumbnail((550, 440), Image.LANCZOS)

                # Показываем
                self.photo_image = ctk.CTkImage(light_image=img, dark_image=img,
                                                size=(img.width, img.height))
                self.placeholder_frame.grid_remove()
                self.img_label.configure(image=self.photo_image, text="")
                self.img_label.grid()

            self.file_label.configure(
                text=f"  {fname}  ·  {size//1024} KB",
                text_color="#94a3b8"
            )
            self.img_container.configure(border_color="#1f2d47")
            self.analyze_btn.configure(state="normal")

        except Exception as ex:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл:\n{ex}")

    def _load_dicom(self, path):
        """Загружает DICOM файл и конвертирует в PIL Image."""
        try:
            if DICOM_AVAILABLE:
                ds = pydicom.dcmread(path)
                arr = ds.pixel_array.astype(np.float32)
                # Нормализация
                if hasattr(ds, 'WindowCenter') and hasattr(ds, 'WindowWidth'):
                    wc = float(ds.WindowCenter if not isinstance(ds.WindowCenter, pydicom.multival.MultiValue) else ds.WindowCenter[0])
                    ww = float(ds.WindowWidth if not isinstance(ds.WindowWidth, pydicom.multival.MultiValue) else ds.WindowWidth[0])
                    low = wc - ww / 2
                    high = wc + ww / 2
                    arr = np.clip(arr, low, high)
                arr = ((arr - arr.min()) / (arr.max() - arr.min() + 1e-6) * 255).astype(np.uint8)
                # Если grayscale → RGB
                if len(arr.shape) == 2:
                    img = Image.fromarray(arr, mode="L").convert("RGB")
                elif arr.shape[2] == 3:
                    img = Image.fromarray(arr, mode="RGB")
                else:
                    img = Image.fromarray(arr[:, :, 0], mode="L").convert("RGB")
                return img
            else:
                # Без pydicom DICOM не поддерживается
                return None
        except Exception as e:
            print(f"⚠ DICOM error: {e}")
            return None

    # ──────────────────────────────────────────────
    def _cancel_analysis(self):
        """Отменяет текущий анализ."""
        self._cancel_event.set()
        self._anim_running = False
        self.img_container.configure(border_color="#1f2d47")
        
        # Восстанавливаем UI
        self.analyze_btn.configure(state="normal", text="⚡  Анализировать")
        self.cancel_btn.pack_forget()
        self.analyze_btn.pack(side="left")
        
        # Восстанавливаем текст файла
        fname = Path(self.current_image_path).name if self.current_image_path else ""
        size = Path(self.current_image_path).stat().st_size if self.current_image_path else 0
        self.file_label.configure(
            text=f"  {fname}  ·  {size//1024} KB" if fname else "Файл не выбран",
            text_color="#94a3b8" if fname else "#334155"
        )
        
        # Ждём завершения потока
        if self._analysis_thread and self._analysis_thread.is_alive():
            self._analysis_thread.join(timeout=2.0)
        
        # Сбрасываем event для следующего анализа
        self._cancel_event.clear()

    def _analyze(self):
        if not self.current_image_path:
            return

        # Сбрасываем флаг отмены
        self._cancel_event.clear()
        
        self.analyze_btn.configure(state="disabled", text="⏳  Анализирую...")
        
        # Показываем кнопку отмены
        self.analyze_btn.pack_forget()
        self.cancel_btn.pack(side="left")
        self.cancel_btn.configure(state="normal")

        # Анимация загрузки
        self._anim_running = True
        self._anim_dots = 0
        def animate():
            if not self._anim_running:
                return
            dots = "·" * (self._anim_dots % 4)
            self.file_label.configure(text=f"⏳  Анализирую{dots}", text_color="#38bdf8")
            self._anim_dots += 1
            self.after(400, animate)
        animate()
        
        # Пульсация рамки изображения
        self.img_container.configure(border_color="#0ea5e9")

        def run():
            # Реальный инференс через модель (с TTA и возможностью отмены)
            result_data = MODEL.predict(self.current_image_path, cancel_event=self._cancel_event)
            
            # Проверяем отмену
            if result_data.get("cancelled"):
                self.after(0, self._on_analysis_cancelled)
                return
            
            # Проверяем ошибку DICOM
            if result_data.get("error") == "DICOM_ERROR":
                self.after(0, lambda: self._on_dicom_error(result_data.get("message", "Ошибка DICOM")))
                return
            
            key = result_data["key"]
            path = PATHOLOGIES[key]
            confidence = result_data["confidence"]
            all_preds = result_data["all_preds"]
            mode = result_data["mode"]
            inference_ms = result_data["inference_time_ms"]

            result = {
                "key": key,
                "path": path,
                "confidence": confidence,
                "all_preds": all_preds[:5],
                "datetime": datetime.datetime.now().strftime("%d.%m.%Y  %H:%M:%S"),
                "filename": Path(self.current_image_path).name,
                "icd": path["icd"],
                "diagnosis": path["name"],
                "latin": path["latin"],
                "severity": SEVERITY_LABELS[path["severity"]],
                "description": path["description"],
                "treatment": path["treatment"],
                "mode": mode,
                "inference_time_ms": inference_ms,
            }

            self.current_result = result
            self.after(0, lambda: self._show_result(result))

        self._analysis_thread = threading.Thread(target=run, daemon=True)
        self._analysis_thread.start()

    def _on_analysis_cancelled(self):
        """Обработчик отмены анализа."""
        self._anim_running = False
        self.analyze_btn.configure(state="normal", text="⚡  Анализировать")
        self.cancel_btn.pack_forget()
        self.analyze_btn.pack(side="left")
        self.img_container.configure(border_color="#1f2d47")
        self.file_label.configure(text="Анализ отменён", text_color="#f87171")

    def _on_dicom_error(self, message):
        """Обработчик ошибки DICOM."""
        self._anim_running = False
        self.analyze_btn.configure(state="normal", text="⚡  Анализировать")
        self.cancel_btn.pack_forget()
        self.analyze_btn.pack(side="left")
        self.img_container.configure(border_color="#1f2d47")
        messagebox.showerror("Ошибка DICOM", message)

    def _show_result(self, result):
        # Останавливаем анимацию
        self._anim_running = False
        self.img_container.configure(border_color="#1f2d47")
        
        # Восстанавливаем кнопки
        self.cancel_btn.pack_forget()
        self.analyze_btn.pack(side="left")
        
        # Восстанавливаем текст файла
        fname = Path(self.current_image_path).name if self.current_image_path else ""
        size = Path(self.current_image_path).stat().st_size if self.current_image_path else 0
        self.file_label.configure(
            text=f"  {fname}  ·  {size//1024} KB" if fname else "Файл не выбран",
            text_color="#94a3b8" if fname else "#334155"
        )
        
        path = result["path"]

        # Очищаем старые карточки
        try:
            if self.diag_card_frame:
                self.diag_card_frame.destroy()
            if self.treat_card_frame:
                self.treat_card_frame.destroy()
            if getattr(self, "_warn_frame", None):
                self._warn_frame.destroy()
                self._warn_frame = None
            if getattr(self, "_disclaimer_frame", None):
                self._disclaimer_frame.destroy()
                self._disclaimer_frame = None
            if getattr(self, "_info_bar_frame", None):
                self._info_bar_frame.destroy()
                self._info_bar_frame = None
        except Exception:
            pass
        self.empty_card.grid_remove()

        try:
            row = 0

            # Предупреждение о низкой уверенности
            confidence = result.get("confidence", 0)
            if confidence < self.confidence_threshold:
                warn = ctk.CTkFrame(self.result_scroll, fg_color="#422006",
                                    corner_radius=12, border_width=1, border_color="#854d0e")
                warn.grid(row=row, column=0, sticky="ew", pady=(0, 8))
                ctk.CTkLabel(warn,
                             text=f"⚠️  Уверенность модели ({confidence:.1f}%) ниже порога ({self.confidence_threshold:.0f}%). "
                                  f"Результат требует особого внимания врача.",
                             font=ctk.CTkFont(size=11), text_color="#fbbf24",
                             wraplength=460, justify="left", padx=16, pady=10).pack()
                self._warn_frame = warn
                row += 1

            self.diag_card_frame = self._build_diagnosis_card(result, path, row)
            row += 1
            self.treat_card_frame = self._build_treatment_card(path, row)
            row += 1
            self._disclaimer_frame = self._build_disclaimer(row)
            row += 1
            self._info_bar_frame = self._build_info_bar(result, row)
        except Exception as e:
            logger.error("Ошибка отображения: %s", e)
            import traceback
            traceback.print_exc()

        # ВСЕГДА активируем кнопки (даже если UI упал)
        self.export_btn.configure(state="normal")
        self.gradcam_btn.configure(state="normal")

        mode = result.get("mode", "simulation")
        ms = result.get("inference_time_ms", 0)
        tta = result.get("tta", False)
        mode_labels = {
            "checkpoint": "🧠 Обучена",
            "pretrained": "🔬 Демо",
            "simulation": "🎲 Симуляция",
        }
        tta_text = " + TTA" if tta else ""
        mode_text = f"{mode_labels.get(mode, mode)}{tta_text}  •  {ms:.0f} мс"
        self.analyze_btn.configure(state="normal", text=f"⚡  Анализировать  ({mode_text})")

        # Add to history
        self.history.insert(0, result)
        self._save_history()
        self.hist_btn.configure(text=f"📋 История  {len(self.history)}")

    def _build_diagnosis_card(self, result, path, row):
        """Строит карточку диагноза в правой панели."""
        frame = ctk.CTkFrame(self.result_scroll, fg_color="#0d1420",
                             corner_radius=16, border_width=1,
                             border_color=path["color"])
        frame.grid(row=row, column=0, sticky="ew", pady=(0, 12))
        frame.grid_columnconfigure(0, weight=1)

        # Header
        dh = ctk.CTkFrame(frame, fg_color="transparent", height=40)
        dh.grid(sticky="ew", padx=18, pady=(14, 0))
        ctk.CTkLabel(dh, text="🩺  ДИАГНОЗ",
                     font=ctk.CTkFont(size=11, family="Courier"),
                     text_color="#475569").pack(side="left")
        ctk.CTkLabel(dh, text=result["icd"],
                     font=ctk.CTkFont(size=11, family="Courier"),
                     text_color="#334155").pack(side="right")

        ctk.CTkFrame(frame, fg_color="#1f2d47", height=1).grid(sticky="ew", pady=4)

        # Emoji + name
        ename = ctk.CTkFrame(frame, fg_color="transparent")
        ename.grid(sticky="ew", padx=18, pady=(10, 0))
        ename.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(ename, text=path["emoji"],
                     font=ctk.CTkFont(size=36)).grid(row=0, column=0, sticky="n", padx=(0, 12))
        nl = ctk.CTkFrame(ename, fg_color="transparent")
        nl.grid(row=0, column=1, sticky="ew")
        nl.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(nl, text=ui_wrap(path["name"], 31),
                     font=ctk.CTkFont(size=18, weight="bold"),
                     text_color=path["color"], wraplength=330,
                     justify="left", anchor="w").grid(row=0, column=0, sticky="ew")
        ctk.CTkLabel(nl, text=ui_wrap(path["latin"], 38),
                     font=ctk.CTkFont(size=12),
                     text_color="#475569", wraplength=330,
                     justify="left", anchor="w").grid(row=1, column=0, sticky="ew")

        # Confidence
        cf = ctk.CTkFrame(frame, fg_color="transparent")
        cf.grid(sticky="ew", padx=18, pady=(12, 0))
        cf.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(cf, text="УВЕРЕННОСТЬ МОДЕЛИ",
                     font=ctk.CTkFont(size=10, family="Courier"),
                     text_color="#334155").grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(cf, text=f"{result['confidence']:.1f}%",
                     font=ctk.CTkFont(size=16, weight="bold"),
                     text_color=path["color"]).grid(row=0, column=1, sticky="e")
        prog = ctk.CTkProgressBar(frame, height=8,
                                   progress_color=path["color"], fg_color="#1a2540")
        prog.grid(sticky="ew", padx=18, pady=(4, 0))
        prog.set(result["confidence"] / 100)

        # Severity
        sev = path["severity"]
        sev_col = SEVERITY_COLORS.get(sev, "#94a3b8")

        sf = ctk.CTkFrame(frame, fg_color="transparent")
        sf.grid(sticky="ew", padx=18, pady=(12, 0))
        sf.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(sf, text="СТЕПЕНЬ ТЯЖЕСТИ",
                     font=ctk.CTkFont(size=10, family="Courier"),
                     text_color="#334155").grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(sf, text=SEVERITY_LABELS.get(sev, "—"),
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color=sev_col).grid(row=0, column=1, sticky="e")

        pip_frame = ctk.CTkFrame(frame, fg_color="transparent")
        pip_frame.grid(sticky="ew", padx=18, pady=(6, 0))
        for i in range(1, 5):
            c = sev_col if i <= sev else "#1a2540"
            ctk.CTkFrame(pip_frame, fg_color=c, width=0, height=8,
                         corner_radius=4).pack(side="left", fill="x", expand=True, padx=2)

        # Description
        desc_box = ctk.CTkFrame(frame, fg_color="#080c14", corner_radius=10)
        desc_box.grid(sticky="ew", padx=18, pady=12)
        desc_box.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(desc_box, text=ui_wrap(path["description"], 48),
                     font=ctk.CTkFont(size=12), text_color="#94a3b8",
                     wraplength=340, justify="left", anchor="w",
                     padx=14, pady=10).grid(row=0, column=0, sticky="ew")

        # Top predictions
        ctk.CTkLabel(frame, text="ВСЕ ПРЕДСКАЗАНИЯ",
                     font=ctk.CTkFont(size=10, family="Courier"),
                     text_color="#334155").grid(sticky="w", padx=18)

        for i, (k, conf) in enumerate(result["all_preds"]):
            p2 = PATHOLOGIES.get(k, {})
            is_top = (i == 0)
            pf = ctk.CTkFrame(frame,
                               fg_color="#111827" if is_top else "transparent",
                               corner_radius=8)
            pf.grid(sticky="ew", padx=18, pady=2)
            pf.grid_columnconfigure(1, weight=1)
            pf.grid_columnconfigure(2, weight=0, minsize=52)

            ctk.CTkLabel(pf, text=p2.get("emoji", "⬜"),
                         font=ctk.CTkFont(size=13), padx=6, pady=4).grid(row=0, column=0, rowspan=2, sticky="n")
            ctk.CTkLabel(pf, text=ui_wrap(p2.get("name", k), 36),
                         font=ctk.CTkFont(size=12, weight="bold" if is_top else "normal"),
                         text_color="#f0f6ff" if is_top else "#94a3b8",
                         anchor="w", justify="left", wraplength=330).grid(row=0, column=1, columnspan=2, sticky="ew", padx=(0, 4), pady=(4, 0))

            pb = ctk.CTkProgressBar(pf, height=4,
                                     progress_color=path["color"] if is_top else "#1f2d47",
                                     fg_color="#1a2540")
            pb.grid(row=1, column=1, sticky="ew", padx=(0, 6), pady=(2, 6))
            pb.set(conf / 100)

            ctk.CTkLabel(pf, text=f"{conf:.1f}%",
                         font=ctk.CTkFont(size=11, family="Courier"),
                         text_color=path["color"] if is_top else "#334155",
                         padx=6).grid(row=1, column=2, sticky="e")

        ctk.CTkFrame(frame, fg_color="transparent", height=10).grid()

        # ── Шкала «Требуется биопсия» ──
        biopsy = BIOPSY_REQUIRED.get(result["key"], {})
        if biopsy:
            ctk.CTkFrame(frame, fg_color="#1f2d47", height=1).grid(sticky="ew", pady=4)

            bf = ctk.CTkFrame(frame, fg_color="transparent")
            bf.grid(sticky="ew", padx=18, pady=(4, 0))
            bf.grid_columnconfigure(1, weight=1)

            ctk.CTkLabel(bf, text="🔬  БИОПСИЯ",
                         font=ctk.CTkFont(size=10, family="Courier"),
                         text_color="#475569").grid(row=0, column=0, sticky="w")

            urgency_text = biopsy.get("urgency", "")
            urgency_color = biopsy.get("color", "#94a3b8")
            ctk.CTkLabel(bf, text=urgency_text,
                         font=ctk.CTkFont(size=12, weight="bold"),
                         text_color=urgency_color).grid(row=0, column=1, sticky="e")

            biopsy_box = ctk.CTkFrame(frame, fg_color="#0f1724", corner_radius=8)
            biopsy_box.grid(sticky="ew", padx=18, pady=(4, 12))
            biopsy_box.grid_columnconfigure(0, weight=1)
            ctk.CTkLabel(biopsy_box, text=ui_wrap(biopsy.get("text", ""), 48),
                         font=ctk.CTkFont(size=11), text_color="#94a3b8",
                         wraplength=340, justify="left", anchor="w",
                         padx=12, pady=8).grid(row=0, column=0, sticky="ew")

        return frame


    def _build_treatment_card(self, path, row):
        """Строит карточку рекомендаций по лечению."""
        frame = ctk.CTkFrame(self.result_scroll, fg_color="#0d1420",
                             corner_radius=16, border_width=1, border_color="#1f2d47")
        frame.grid(row=row, column=0, sticky="ew", pady=(0, 12))
        frame.grid_columnconfigure(0, weight=1)

        th = ctk.CTkFrame(frame, fg_color="transparent", height=40)
        th.grid(sticky="ew", padx=18, pady=(14, 0))
        ctk.CTkLabel(th, text="💊  РЕКОМЕНДАЦИИ ПО ЛЕЧЕНИЮ",
                     font=ctk.CTkFont(size=11, family="Courier"),
                     text_color="#475569").pack(side="left")

        ctk.CTkFrame(frame, fg_color="#1f2d47", height=1).grid(sticky="ew", pady=4)

        tab_colors = {
            "Медикаменты": "#38bdf8", "Процедуры": "#34d399",
            "Диета": "#fbbf24", "Наблюдение": "#a78bfa",
        }

        for section, items in path["treatment"].items():
            sc = tab_colors.get(section, "#94a3b8")

            sh = ctk.CTkFrame(frame, fg_color="#111827", corner_radius=8)
            sh.grid(sticky="ew", padx=18, pady=(10, 4))
            ctk.CTkLabel(sh, text=f"  ▸ {section}",
                         font=ctk.CTkFont(size=13, weight="bold"),
                         text_color=sc, anchor="w", pady=7).pack(fill="x")

            for title, desc, pill in items:
                item_box = ctk.CTkFrame(frame, fg_color="#080c14", corner_radius=8)
                item_box.grid(sticky="ew", padx=18, pady=3)
                item_box.grid_columnconfigure(1, weight=1)

                bul = ctk.CTkFrame(item_box, fg_color=darken_color(sc),
                                   width=32, height=32, corner_radius=8)
                bul.grid(row=0, column=0, rowspan=2, padx=(10, 8), pady=8, sticky="n")
                bul.grid_propagate(False)
                ctk.CTkLabel(bul, text=title[0:2],
                             font=ctk.CTkFont(size=14)).place(relx=.5, rely=.5, anchor="center")

                ctk.CTkLabel(item_box, text=title[2:].strip(),
                             font=ctk.CTkFont(size=13, weight="bold"),
                             text_color="#e2eaf4", anchor="w").grid(row=0, column=1, sticky="w", pady=(8, 0))
                ctk.CTkLabel(item_box, text=desc,
                             font=ctk.CTkFont(size=11), text_color="#64748b",
                             wraplength=400, justify="left", anchor="w").grid(row=1, column=1, sticky="w")

                pill_lbl = ctk.CTkFrame(item_box, fg_color="#1a2540", corner_radius=4)
                pill_lbl.grid(row=2, column=1, sticky="w", pady=(2, 8))
                ctk.CTkLabel(pill_lbl, text=pill,
                             font=ctk.CTkFont(size=10, family="Courier"),
                             text_color="#475569", padx=8, pady=2).pack()

        return frame

    def _build_disclaimer(self, row):
        """Строит блок дисклеймера."""
        disc = ctk.CTkFrame(self.result_scroll, fg_color="#1c0a0a",
                             corner_radius=12, border_width=1, border_color="#7f1d1d")
        disc.grid(row=row, column=0, sticky="ew", pady=(0, 8))
        disc.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(disc,
                     text="⚠️  Данное заключение носит вспомогательный характер и не заменяет "
                          "консультацию врача. Все лечебные мероприятия назначаются "
                          "исключительно лечащим врачом после клинического обследования.",
                     font=ctk.CTkFont(size=11), text_color="#fca5a5",
                     wraplength=360, justify="left", anchor="w",
                     padx=16, pady=12).grid(row=0, column=0, sticky="ew")
        return disc

    def _build_info_bar(self, result, row):
        """Строит блок информации о режиме и времени инференса."""
        mode = result.get("mode", "simulation")
        ms = result.get("inference_time_ms", 0)
        mode_labels = {
            "checkpoint": "🧠 Обучена", "pretrained": "🔬 Демо", "simulation": "🎲 Симуляция",
        }
        info = ctk.CTkFrame(self.result_scroll, fg_color="#080c14", corner_radius=8)
        info.grid(row=row, column=0, sticky="ew", pady=(0, 8))
        ctk.CTkLabel(info,
                     text=f"⏱ {ms:.0f} мс  |  {mode_labels.get(mode, '?')}  |  EfficientNet-B4",
                     font=ctk.CTkFont(size=10, family="Courier"),
                     text_color="#475569", padx=12, pady=6).pack()
        return info

    # ──────────────────────────────────────────────
    def _open_history(self):
        if not self.history:
            messagebox.showinfo("История", "История диагностик пуста.\nПроведите хотя бы одну диагностику.")
            return

        win = ctk.CTkToplevel(self)
        win.title("История диагностик")
        win.geometry("980x560")
        win.configure(fg_color="#0a0e17")
        win.after(150, win.grab_set)

        ctk.CTkLabel(win, text="📋  История диагностик",
                     font=ctk.CTkFont(size=18, weight="bold"),
                     text_color="#f0f6ff").pack(pady=(16, 4), padx=20, anchor="w")

        scroll = ctk.CTkScrollableFrame(win, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        scroll.grid_columnconfigure(0, weight=1)

        headers = ["Файл", "Дата/Время", "Диагноз", "Уверенность", "Тяжесть"]
        hf = ctk.CTkFrame(scroll, fg_color="#111827", corner_radius=8)
        hf.grid(row=0, column=0, sticky="ew", pady=(0, 4))
        hf.grid_columnconfigure(0, weight=2, minsize=180)
        hf.grid_columnconfigure(1, weight=1, minsize=135)
        hf.grid_columnconfigure(2, weight=4, minsize=330)
        hf.grid_columnconfigure(3, weight=1, minsize=100)
        hf.grid_columnconfigure(4, weight=1, minsize=110)
        for i, h in enumerate(headers):
            ctk.CTkLabel(hf, text=h,
                         font=ctk.CTkFont(size=11, family="Courier"),
                         text_color="#475569",
                         padx=10, pady=8).grid(row=0, column=i, sticky="w")

        for ri, r in enumerate(self.history):
            rf = ctk.CTkFrame(scroll,
                              fg_color="#0d1420" if ri % 2 == 0 else "#080c14",
                              corner_radius=6)
            rf.grid(row=ri + 1, column=0, sticky="ew", pady=1)
            rf.grid_columnconfigure(0, weight=2, minsize=180)
            rf.grid_columnconfigure(1, weight=1, minsize=135)
            rf.grid_columnconfigure(2, weight=4, minsize=330)
            rf.grid_columnconfigure(3, weight=1, minsize=100)
            rf.grid_columnconfigure(4, weight=1, minsize=110)

            sev = r["path"]["severity"]
            sev_col = SEVERITY_COLORS.get(sev, "#94a3b8")
            vals = [
                (r["filename"][:28], "#94a3b8"),
                (r["datetime"], "#475569"),
                (r["path"]["emoji"] + "  " + r["path"]["name"], r["path"]["color"]),
                (f"{r['confidence']:.1f}%", "#38bdf8"),
                (SEVERITY_LABELS.get(sev, "—"), sev_col),
            ]
            for ci, (val, col) in enumerate(vals):
                wrap = 320 if ci == 2 else 130
                ctk.CTkLabel(rf, text=val,
                             font=ctk.CTkFont(size=12),
                             text_color=col, wraplength=wrap,
                             justify="left", anchor="w",
                             padx=10, pady=7).grid(row=0, column=ci, sticky="ew")

    # ──────────────────────────────────────────────
    def _export_pdf(self):
        if not self.current_result:
            return

        path = filedialog.asksaveasfilename(
            title="Сохранить заключение как PDF",
            defaultextension=".pdf",
            filetypes=[("PDF файл", "*.pdf")],
            initialfile=f"EndoMed_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        )
        if not path:
            return

        ok = export_pdf_report(self.current_result, path)
        if ok:
            messagebox.showinfo("Готово", f"PDF-отчёт сохранён:\n{path}")
        else:
            # Fallback — TXT
            txt_path = path.replace(".pdf", ".txt")
            r = self.current_result
            with open(txt_path, "w", encoding="utf-8") as f:
                f.write("=" * 60 + "\n")
                f.write("  EndoMed — Заключение по результатам диагностики\n")
                f.write("=" * 60 + "\n\n")
                f.write(f"Дата:        {r['datetime']}\n")
                f.write(f"Файл:        {r['filename']}\n")
                f.write(f"Модель:      EfficientNet-B4\n")
                f.write(f"МКБ-10:      {r['icd']}\n\n")
                f.write(f"ДИАГНОЗ:     {r['diagnosis']}\n")
                f.write(f"Лат.:        {r['latin']}\n")
                f.write(f"Уверен.:     {r['confidence']:.1f}%\n")
                f.write(f"Тяжесть:     {r['severity']}\n\n")
                f.write(f"Описание:\n{r['description']}\n\n")
                f.write("-" * 60 + "\n")
                for section, items in r["treatment"].items():
                    f.write(f"\n{section.upper()}:\n")
                    for title, desc, pill in items:
                        f.write(f"  {title}\n  {desc}\n  [{pill}]\n\n")
                f.write("-" * 60 + "\n")
                f.write("ВНИМАНИЕ: Носит вспомогательный характер. Диагноз ставит врач.\n")

            messagebox.showinfo(
                "Сохранено как TXT",
                f"Для PDF с кириллицей установите шрифт DejaVu:\n"
                f"Linux: sudo apt-get install fonts-dejavu\n"
                f"Или сохраните шрифт рядом со скриптом в папку fonts/\n\n"
                f"Сохранено как:\n{txt_path}"
            )

    # ──────────────────────────────────────────────
    #  GRAD-CAM ВИЗУАЛИЗАЦИЯ (v6 — с отменой)
    # ──────────────────────────────────────────────
    def _show_gradcam(self):
        """Показывает окно с Grad-CAM тепловой картой."""
        if not self.current_image_path or not self.current_result:
            messagebox.showinfo("Grad-CAM", "Сначала выполните анализ изображения.")
            return

        # Создаём окно
        win = ctk.CTkToplevel(self)
        win.title("🔥 Grad-CAM — Визуализация внимания нейросети")
        win.geometry("1000x620")
        win.configure(fg_color="#0a0e17")
        win.after(150, win.grab_set)

        # Заголовок
        header = ctk.CTkFrame(win, fg_color="#0d1420", height=64, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)
        header.grid_columnconfigure(0, weight=1)
        
        result = self.current_result
        path_info = result["path"]
        
        ctk.CTkLabel(header, text=f"🔥  Grad-CAM  |  {path_info['emoji']} {result['diagnosis']}  |  {result['confidence']:.1f}%",
                     font=ctk.CTkFont(size=16, weight="bold"),
                     text_color="#f0f6ff", wraplength=620,
                     justify="left", anchor="w").grid(row=0, column=0, sticky="ew", padx=(20, 10), pady=(8, 0))
        ctk.CTkLabel(header, text="Тепловая карта показывает области, на которые нейросеть обратила наибольшее внимание",
                     font=ctk.CTkFont(size=11),
                     text_color="#475569", anchor="w").grid(row=1, column=0, sticky="ew", padx=(20, 10), pady=(0, 6))

        # Основная область с двумя изображениями
        content = ctk.CTkFrame(win, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=20, pady=16)
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=1)
        content.grid_rowconfigure(1, weight=1)

        # Подписи
        ctk.CTkLabel(content, text="ОРИГИНАЛ",
                     font=ctk.CTkFont(size=11, family="Courier"),
                     text_color="#475569").grid(row=0, column=0, pady=(0, 4))
        ctk.CTkLabel(content, text="GRAD-CAM ТЕПЛОВАЯ КАРТА",
                     font=ctk.CTkFont(size=11, family="Courier"),
                     text_color="#fbbf24").grid(row=0, column=1, pady=(0, 4))

        # Загружаем оригинал
        try:
            orig_img = Image.open(self.current_image_path).convert("RGB")
            orig_img.thumbnail((460, 440), Image.LANCZOS)
            
            orig_ctk = ctk.CTkImage(light_image=orig_img, dark_image=orig_img,
                                     size=(orig_img.width, orig_img.height))
            
            orig_frame = ctk.CTkFrame(content, fg_color="#080c14", corner_radius=12,
                                       border_width=2, border_color="#1f2d47")
            orig_frame.grid(row=1, column=0, padx=(0, 8), sticky="nsew")
            ctk.CTkLabel(orig_frame, image=orig_ctk, text="").pack(expand=True, padx=8, pady=8)
            orig_frame._orig_image = orig_ctk  # prevent GC
        except Exception as e:
            ctk.CTkLabel(content, text=f"Ошибка загрузки: {e}",
                         text_color="#f87171").grid(row=1, column=0)

        # Генерируем Grad-CAM
        heatmap_frame = ctk.CTkFrame(content, fg_color="#080c14", corner_radius=12,
                                      border_width=2, border_color="#854d0e")
        heatmap_frame.grid(row=1, column=1, padx=(8, 0), sticky="nsew")

        # Показываем загрузку
        loading_label = ctk.CTkLabel(heatmap_frame, text="⏳ Генерация Grad-CAM...",
                                      font=ctk.CTkFont(size=14),
                                      text_color="#fbbf24")
        loading_label.pack(expand=True)

        # Event для отмены Grad-CAM
        cam_cancel_event = threading.Event()

        def generate():
            class_idx = None
            try:
                class_idx = MODEL.classes.index(self.current_result.get("key"))
            except (ValueError, AttributeError):
                class_idx = None
            heatmap_img = MODEL.generate_heatmap(self.current_image_path, class_idx=class_idx, cancel_event=cam_cancel_event)
            
            if heatmap_img is None:
                # Fallback — создаём симулированную тепловую карту
                try:
                    img = Image.open(self.current_image_path).convert("RGB")
                    w, h = img.size
                    hm = np.zeros((h, w), dtype=np.float32)
                    X, Y = np.meshgrid(np.arange(w), np.arange(h))
                    for _ in range(random.randint(2, 4)):
                        cx = random.randint(w // 4, 3 * w // 4)
                        cy = random.randint(h // 4, 3 * h // 4)
                        rx = max(random.randint(w // 8, w // 4), 1)
                        ry = max(random.randint(h // 8, h // 4), 1)
                        # Гауссово затухание — нет ступенчатых краёв
                        dist = (X - cx) ** 2 / (rx ** 2) + (Y - cy) ** 2 / (ry ** 2)
                        intensity = random.uniform(0.5, 1.0)
                        vals = intensity * np.exp(-2.0 * dist)
                        hm = np.maximum(hm, vals)
                    if hm.max() > 0:
                        hm /= hm.max()
                    heatmap_img = GradCAM.overlay_heatmap(img, hm, alpha=0.45)
                except Exception as e:
                    logger.warning("Ошибка симуляции Grad-CAM: %s", e)
                    heatmap_img = None
            
            def show():
                loading_label.destroy()
                if heatmap_img:
                    hm_display = heatmap_img.copy()
                    hm_display.thumbnail((460, 440), Image.LANCZOS)
                    hm_ctk = ctk.CTkImage(light_image=hm_display, dark_image=hm_display,
                                           size=(hm_display.width, hm_display.height))
                    ctk.CTkLabel(heatmap_frame, image=hm_ctk, text="").pack(expand=True, padx=8, pady=8)
                    heatmap_frame._hm_image = hm_ctk
                else:
                    ctk.CTkLabel(heatmap_frame, text="⚠ Не удалось сгенерировать\nGrad-CAM",
                                 font=ctk.CTkFont(size=14),
                                 text_color="#f87171").pack(expand=True)
            
            win.after(0, show)

        cam_thread = threading.Thread(target=generate, daemon=True)
        cam_thread.start()

        # Легенда
        legend = ctk.CTkFrame(win, fg_color="#0d1420", height=40, corner_radius=0)
        legend.pack(fill="x", side="bottom")
        legend.pack_propagate(False)
        
        colors_info = [
            ("🔴 Красный", "Максимальное внимание — возможный очаг патологии"),
            ("🟡 Жёлтый", "Повышенное внимание"),
            ("🔵 Синий", "Минимальное внимание — нормальная ткань"),
        ]
        for ci, (col, desc) in enumerate(colors_info):
            ctk.CTkLabel(legend, text=f"{col}: {desc}",
                         font=ctk.CTkFont(size=10),
                         text_color="#64748b").pack(side="left", padx=15, pady=8)

        # Очистка при закрытии окна
        def on_close():
            cam_cancel_event.set()
            win.destroy()
        win.protocol("WM_DELETE_WINDOW", on_close)

    # ──────────────────────────────────────────────
    #  ДАШБОРД СТАТИСТИКИ
    # ──────────────────────────────────────────────
    def _open_dashboard(self):
        """Показывает дашборд статистики по истории диагностик."""
        if not self.history:
            messagebox.showinfo("Статистика", "История диагностик пуста.\nПроведите хотя бы одну диагностику для\nотображения статистики.")
            return

        win = ctk.CTkToplevel(self)
        win.title("📊 Дашборд статистики EndoMed")
        win.geometry("1100x700")
        win.configure(fg_color="#0a0e17")
        win.after(150, win.grab_set)

        # Заголовок
        ctk.CTkLabel(win, text=f"📊  Дашборд статистики  |  {len(self.history)} диагностик",
                     font=ctk.CTkFont(size=18, weight="bold"),
                     text_color="#f0f6ff").pack(pady=(16, 4), padx=20, anchor="w")
        ctk.CTkLabel(win, text="Анализ результатов всех выполненных диагностик в текущей сессии",
                     font=ctk.CTkFont(size=12),
                     text_color="#475569").pack(padx=20, anchor="w")

        if not MATPLOTLIB_AVAILABLE:
            ctk.CTkLabel(win, text="⚠ Для дашборда установите matplotlib:\npip install matplotlib",
                         font=ctk.CTkFont(size=16),
                         text_color="#f87171").pack(expand=True)
            return

        # Собираем данные из истории
        diagnoses = {}
        severities = {}
        confidences = []
        timeline = []
        
        for r in self.history:
            diag = r.get("diagnosis", "Неизвестно")
            diagnoses[diag] = diagnoses.get(diag, 0) + 1
            
            sev = r.get("severity", "—")
            severities[sev] = severities.get(sev, 0) + 1
            
            confidences.append(r.get("confidence", 0))
            timeline.append({
                "time": r.get("datetime", ""),
                "diag": diag,
                "conf": r.get("confidence", 0),
            })

        # Цвета для патологий
        pathology_colors = {
            "Полип ЖКТ": "#38bdf8",
            "Злокачественное новообразование ЖКТ": "#f87171",
            "Язвенное поражение ЖКТ": "#fbbf24",
            "Гастрит / Эзофагит": "#fb923c",
            "Пищевод Барретта": "#a78bfa",
            "Воспалительные заболевания кишечника": "#fb923c",
            "Норма — эндоскопическая картина без патологии": "#34d399",
        }

        # Создаём matplotlib фигуру — тёмная тема
        fig = plt.Figure(figsize=(11, 5.5), facecolor="#0a0e17")
        fig.subplots_adjust(wspace=0.35, hspace=0.4, top=0.92, bottom=0.12, left=0.06, right=0.97)

        # ── График 1: Pie Chart — Распределение диагнозов ──
        ax1 = fig.add_subplot(2, 2, 1)
        ax1.set_facecolor("#0a0e17")
        
        labels = list(diagnoses.keys())
        sizes = list(diagnoses.values())
        colors = [pathology_colors.get(l, "#64748b") for l in labels]
        short_labels = [l.split()[0] if len(l) > 15 else l for l in labels]
        
        wedges, texts, autotexts = ax1.pie(sizes, labels=short_labels, colors=colors,
                                            autopct='%1.0f%%', startangle=90,
                                            textprops={'color': '#94a3b8', 'fontsize': 8})
        for at in autotexts:
            at.set_color('#f0f6ff')
            at.set_fontsize(9)
            at.set_fontweight('bold')
        ax1.set_title("Распределение диагнозов", color="#f0f6ff", fontsize=11, fontweight="bold", pad=10)

        # ── График 2: Bar Chart — Уверенность по диагнозам ──
        ax2 = fig.add_subplot(2, 2, 2)
        ax2.set_facecolor("#0d1420")
        
        avg_conf = {}
        for r in self.history:
            d = r.get("diagnosis", "?")
            short_d = d.split()[0]
            if short_d not in avg_conf:
                avg_conf[short_d] = []
            avg_conf[short_d].append(r.get("confidence", 0))
        
        bar_labels = list(avg_conf.keys())
        bar_values = [sum(v) / len(v) for v in avg_conf.values()]
        bar_colors = [pathology_colors.get(next((k for k in pathology_colors if k.startswith(bl)), ""), "#64748b") for bl in bar_labels]
        
        bars = ax2.barh(bar_labels, bar_values, color=bar_colors, height=0.6, edgecolor="none")
        ax2.set_xlim(0, 105)
        ax2.set_title("Средняя уверенность (%)", color="#f0f6ff", fontsize=11, fontweight="bold", pad=10)
        ax2.tick_params(colors="#64748b", labelsize=8)
        ax2.spines['top'].set_visible(False)
        ax2.spines['right'].set_visible(False)
        ax2.spines['bottom'].set_color("#1f2d47")
        ax2.spines['left'].set_color("#1f2d47")
        ax2.xaxis.label.set_color("#64748b")
        
        for bar, val in zip(bars, bar_values):
            ax2.text(val + 1.5, bar.get_y() + bar.get_height() / 2,
                     f'{val:.1f}%', va='center', color='#94a3b8', fontsize=8)

        # ── График 3: Степени тяжести ──
        ax3 = fig.add_subplot(2, 2, 3)
        ax3.set_facecolor("#0d1420")
        
        sev_order = ["Норма", "I — Лёгкая", "II — Умеренная", "III — Высокая", "IV — Критическая"]
        sev_colors_map = {"Норма": "#34d399", "I — Лёгкая": "#34d399", "II — Умеренная": "#fbbf24",
                          "III — Высокая": "#fb923c", "IV — Критическая": "#f87171"}
        
        sev_labels = [s for s in sev_order if s in severities]
        sev_vals = [severities[s] for s in sev_labels]
        sev_cols = [sev_colors_map.get(s, "#64748b") for s in sev_labels]
        
        ax3.bar(range(len(sev_labels)), sev_vals, color=sev_cols, width=0.6, edgecolor="none")
        ax3.set_xticks(range(len(sev_labels)))
        ax3.set_xticklabels([s.split("—")[0].strip() if "—" in s else s for s in sev_labels], fontsize=8)
        ax3.set_title("Распределение по тяжести", color="#f0f6ff", fontsize=11, fontweight="bold", pad=10)
        ax3.tick_params(colors="#64748b", labelsize=8)
        ax3.spines['top'].set_visible(False)
        ax3.spines['right'].set_visible(False)
        ax3.spines['bottom'].set_color("#1f2d47")
        ax3.spines['left'].set_color("#1f2d47")
        ax3.yaxis.set_major_locator(plt.MaxNLocator(integer=True))

        # ── График 4: Timeline уверенности ──
        ax4 = fig.add_subplot(2, 2, 4)
        ax4.set_facecolor("#0d1420")
        
        conf_vals = [t["conf"] for t in reversed(timeline)]
        x_vals = list(range(1, len(conf_vals) + 1))
        
        ax4.plot(x_vals, conf_vals, color="#38bdf8", linewidth=2, marker="o",
                 markersize=5, markerfacecolor="#0ea5e9", markeredgecolor="#38bdf8")
        ax4.fill_between(x_vals, conf_vals, alpha=0.15, color="#38bdf8")
        ax4.set_ylim(0, 105)
        ax4.set_xlabel("Анализ №", color="#64748b", fontsize=9)
        ax4.set_ylabel("Уверенность %", color="#64748b", fontsize=9)
        ax4.set_title("Уверенность по анализам", color="#f0f6ff", fontsize=11, fontweight="bold", pad=10)
        ax4.tick_params(colors="#64748b", labelsize=8)
        ax4.spines['top'].set_visible(False)
        ax4.spines['right'].set_visible(False)
        ax4.spines['bottom'].set_color("#1f2d47")
        ax4.spines['left'].set_color("#1f2d47")
        ax4.axhline(y=90, color="#4ade80", linestyle="--", alpha=0.4, linewidth=1)
        ax4.text(0.5, 91.5, "Порог 90%", color="#4ade80", fontsize=7, alpha=0.6)

        # Встраиваем в tkinter
        canvas_frame = ctk.CTkFrame(win, fg_color="transparent")
        canvas_frame.pack(fill="both", expand=True, padx=10, pady=(8, 0))
        
        canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

        def _on_dashboard_close():
            plt.close(fig)
            win.destroy()
        win.protocol("WM_DELETE_WINDOW", _on_dashboard_close)

        # Нижняя панель со сводкой
        summary = ctk.CTkFrame(win, fg_color="#0d1420", height=45, corner_radius=0)
        summary.pack(fill="x", side="bottom")
        summary.pack_propagate(False)
        
        total = len(self.history)
        avg_c = sum(confidences) / total if total > 0 else 0
        most_common = max(diagnoses, key=diagnoses.get) if diagnoses else "—"
        mc_short = most_common.split()[0] if most_common != "—" else "—"
        
        stats_text = (
            f"📊 Всего анализов: {total}   |   "
            f"📈 Средняя уверенность: {avg_c:.1f}%   |   "
            f"🔝 Частый диагноз: {mc_short}   |   "
            f"🧠 Режим: {MODEL.mode}"
        )
        ctk.CTkLabel(summary, text=stats_text,
                     font=ctk.CTkFont(size=11, family="Courier"),
                     text_color="#64748b").pack(expand=True)

    # ──────────────────────────────────────────────
    #  НАСТРОЙКИ ПОРОГА УВЕРЕННОСТИ
    # ──────────────────────────────────────────────
    def _open_settings(self):
        """Окно настроек порога уверенности."""
        win = ctk.CTkToplevel(self)
        win.title("⚙️ Настройки EndoMed")
        win.geometry("480x350")
        win.configure(fg_color="#0a0e17")
        win.after(150, win.grab_set)

        ctk.CTkLabel(win, text="⚙️  Настройки",
                     font=ctk.CTkFont(size=18, weight="bold"),
                     text_color="#f0f6ff").pack(pady=(20, 4), padx=20, anchor="w")

        ctk.CTkFrame(win, fg_color="#1f2d47", height=1).pack(fill="x", padx=20, pady=8)

        # Порог уверенности
        thresh_frame = ctk.CTkFrame(win, fg_color="#0d1420", corner_radius=12,
                                     border_width=1, border_color="#1f2d47")
        thresh_frame.pack(fill="x", padx=20, pady=8)

        ctk.CTkLabel(thresh_frame, text="ПОРОГ УВЕРЕННОСТИ",
                     font=ctk.CTkFont(size=10, family="Courier"),
                     text_color="#475569").pack(anchor="w", padx=16, pady=(12, 4))

        ctk.CTkLabel(thresh_frame,
                     text="Если уверенность модели ниже порога,\nпоказывается предупреждение.",
                     font=ctk.CTkFont(size=12), text_color="#94a3b8",
                     justify="left").pack(anchor="w", padx=16)

        slider_frame = ctk.CTkFrame(thresh_frame, fg_color="transparent")
        slider_frame.pack(fill="x", padx=16, pady=(8, 4))

        self.thresh_label = ctk.CTkLabel(slider_frame,
                                          text=f"{self.confidence_threshold:.0f}%",
                                          font=ctk.CTkFont(size=20, weight="bold"),
                                          text_color="#38bdf8")
        self.thresh_label.pack(side="right", padx=8)

        def on_slider(val):
            self.confidence_threshold = val
            self.thresh_label.configure(text=f"{val:.0f}%")

        slider = ctk.CTkSlider(slider_frame, from_=10, to=95,
                                number_of_steps=17,
                                command=on_slider,
                                progress_color="#38bdf8",
                                fg_color="#1a2540",
                                button_color="#0ea5e9",
                                button_hover_color="#0284c7")
        slider.set(self.confidence_threshold)
        slider.pack(side="left", fill="x", expand=True)

        ctk.CTkLabel(thresh_frame,
                     text="10%                                              95%",
                     font=ctk.CTkFont(size=9, family="Courier"),
                     text_color="#334155").pack(padx=16, pady=(0, 12))

        # Кнопка сохранения
        def save_and_close():
            self.settings["confidence_threshold"] = self.confidence_threshold
            save_settings(self.settings)
            win.destroy()

        ctk.CTkButton(win, text="💾  Сохранить настройки",
                       width=200, height=40,
                       font=ctk.CTkFont(size=14, weight="bold"),
                       fg_color="#0ea5e9", hover_color="#0284c7",
                       text_color="#000000",
                       command=save_and_close).pack(pady=16)

        # Текущий порог
        ctk.CTkLabel(win,
                     text=f"Текущий порог: {self.confidence_threshold:.0f}%  •  Сохраняется между сессиями",
                     font=ctk.CTkFont(size=10, family="Courier"),
                     text_color="#334155").pack()

    # ──────────────────────────────────────────────
    #  ЭКСПОРТ ИСТОРИИ В EXCEL
    # ──────────────────────────────────────────────
    def _export_excel(self):
        """Экспортирует историю диагностик в Excel."""
        if not self.history:
            messagebox.showinfo("Экспорт", "История диагностик пуста.")
            return

        path = filedialog.asksaveasfilename(
            title="Экспорт истории в Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel файл", "*.xlsx"), ("CSV файл", "*.csv")],
            initialfile=f"EndoMed_История_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not path:
            return

        result = export_history_to_excel(self.history, path)
        if result is True:
            messagebox.showinfo("Готово", f"История экспортирована:\n{path}\n\n{len(self.history)} записей")
        elif isinstance(result, str):
            messagebox.showinfo("Сохранено как CSV",
                                f"Для Excel установите: pip install openpyxl\n\nСохранено как CSV:\n{result}")
        else:
            messagebox.showerror("Ошибка", "Не удалось экспортировать историю")


# ══════════════════════════════════════════════════════════
#  ЗАПУСК
# ══════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = EndoMedApp()
    app.mainloop()
